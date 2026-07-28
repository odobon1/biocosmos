import json
from types import SimpleNamespace

import numpy as np
from openpyxl import load_workbook

from utils.train import ArtifactManager, TrialData, format_mem, merge_mem
from utils.utils import save_pickle, load_pickle, paths


def test_aggregate_metric_stats_keeps_none_leaf_across_trials() -> None:
    # loss_raw["ood"] is never computed -> None in every trial's metrics.json. Aggregating across
    # >1 completed trial must keep it None, not attempt float(None).
    trials = [
        {"scores": {"comp": {"map": {"all": "0.50"}}}, "loss_raw": {"id": "0.7029", "ood": None}, "sim": {"mean": "0.0925"}, "targ": {"mean": "-0.9895"}},
        {"scores": {"comp": {"map": {"all": "0.60"}}}, "loss_raw": {"id": "0.7005", "ood": None}, "sim": {"mean": "0.0935"}, "targ": {"mean": "-0.9885"}},
    ]

    out = ArtifactManager._aggregate_metric_stats(trials, "std")

    assert out["loss_raw"]["ood"] is None
    assert out["loss_raw"]["id"] == "0.7017 ± 0.0017"
    assert out["scores"]["comp"]["map"]["all"] == "55.00 ± 7.07"
    # sim/targ aggregate raw (like loss_raw), not as percentages
    assert out["sim"]["mean"] == "0.0930 ± 0.0007"
    assert out["targ"]["mean"] == "-0.9890 ± 0.0007"


def test_aggregate_metric_stats_ste_spread() -> None:
    # ste = std / sqrt(n): 7.07 / sqrt(2) = 5.00
    trials = [
        {"scores": {"comp": {"map": {"all": "0.50"}}}},
        {"scores": {"comp": {"map": {"all": "0.60"}}}},
    ]

    out = ArtifactManager._aggregate_metric_stats(trials, "ste")

    assert out["scores"]["comp"]["map"]["all"] == "55.00 ± 5.00"


def test_aggregate_metric_stats_single_trial_returns_leaves_verbatim() -> None:
    trials = [{"loss_raw": {"id": "0.7029", "ood": None}}]

    out = ArtifactManager._aggregate_metric_stats(trials, "std")

    assert out == {"loss_raw": {"id": "0.7029", "ood": None}}


def test_update_eval_appends_none_leaves_from_base_eval(tmp_path) -> None:
    # base eval computes no loss -> loss_raw/sim/targ carry None leaves; update_eval must append
    # them as placeholders (not crash on a bare None) so eval curves stay index-aligned across evals
    data = TrialData(tmp_path)
    data.eval_metrics = {
        "scores": {"comp": {"map": {"all": 0.5}}},
        "loss_raw": {"id": None},
        "sim": {"min": None, "max": None, "median": None, "mean": None},
        "targ": {"min": None, "max": None, "median": None, "mean": None},
    }
    data.update_eval(0)
    data.eval_metrics = {
        "scores": {"comp": {"map": {"all": 0.6}}},
        "loss_raw": {"id": 0.7},
        "sim": {"min": -0.02, "max": 0.16, "median": 0.09, "mean": 0.09},
        "targ": {"min": -1.0, "max": 1.0, "median": -1.0, "mean": -0.99},
    }
    data.update_eval(1000)

    assert data.data_eval["n_samps_seen"] == [0, 1000]
    assert data.data_eval["loss_raw"]["id"] == [None, 0.7]
    assert data.data_eval["sim"]["mean"] == [None, 0.09]
    assert data.data_eval["targ"]["min"] == [None, -1.0]


def test_update_metric_stats_counts_trials_lacking_complete_flag(tmp_path, monkeypatch) -> None:
    # completion is now marked by the orchestrator after stats run, so update_metric_stats must aggregate
    # trials by their written final-eval metrics -- not by a `complete` flag that isn't set yet
    dataset = "cub"
    dpath_dataset = tmp_path / dataset
    for seed, all_v in (("42", "0.50"), ("43", "0.60")):
        dpath_final = dpath_dataset / seed / "evals" / "final"
        dpath_final.mkdir(parents=True)
        (dpath_final / "metrics.json").write_text(json.dumps({
            "scores": {"comp": {"map": {"all": all_v}}},
            "loss_raw": {"id": "0.70", "ood": None},
            "sim": {"mean": "0.0925"},
            "targ": {"mean": "-0.9895"},
            "n_samps_seen": "100/100",
        }))

    monkeypatch.setattr(ArtifactManager, "dpath_setting", tmp_path)
    monkeypatch.setattr(ArtifactManager, "dataset", dataset)

    ArtifactManager.update_metric_stats("std")

    stats = json.loads((dpath_dataset / "stats" / "metrics.json").read_text())
    assert stats["n_trials"] == 2
    assert stats["loss_raw"]["ood"] is None
    assert stats["scores"]["comp"]["map"]["all"] == "55.00 ± 7.07"

    listview_text = (dpath_dataset / "stats" / "metrics_listview.json").read_text()
    listview = json.loads(listview_text)
    assert listview["n_trials"] == 2
    assert listview["loss_raw"]["ood"] is None
    assert listview["loss_raw"]["id"] == ["0.7000", "0.7000"]
    assert listview["sim"]["mean"] == ["0.0925", "0.0925"]
    assert listview["scores"]["comp"]["map"]["all"] == ["50.00", "60.00"]
    # each leaf list stays on a single line
    assert '"all": ["50.00", "60.00"]' in listview_text


def _comp(base: float) -> dict:
    return {
        "acc": {"i2t": f"{base + 0.06:.4f}"},
        "map": {
            "all": f"{base:.4f}",
            "ood": f"{base + 0.01:.4f}",
            "id": f"{base + 0.02:.4f}",
            "i2t": f"{base + 0.03:.4f}",
            "i2i": f"{base + 0.04:.4f}",
            "t2i": f"{base + 0.05:.4f}",
        },
    }


def _scores_grp(comp: dict) -> dict:
    # full per-grp scores subtree as written to metrics.json: comp + per-partition primitive scores
    prim = {"map": {"i2t": "0.10", "i2i": "0.10", "t2i": "0.10"}, "acc": {"i2t": "0.10"}}
    return {"comp": comp, "id": prim, "ood": prim}


def test_stats_table_grid_formats_by_trial_count() -> None:
    grid = ArtifactManager._stats_table_grid(
        ("All", "ID", "OOD", "I2T", "I2I", "T2I"),
        [
            ("hp", [_comp(0.50)["map"], _comp(0.60)["map"]]),
            ("sw", [_comp(0.50)["map"]]),
            ("iw", []),
        ],
        "std",
    )

    assert grid[0] == ["Setting", "All", "ID", "OOD", "I2T", "I2I", "T2I"]
    assert [row[0] for row in grid[1:]] == ["hp (2)", "sw (1)", "iw (0)"]
    assert grid[1][1] == "55.00 ± 7.07"  # 2 trials: mean ± std
    assert grid[2][1] == "50.00"  # 1 trial: mean only
    assert grid[3][1] == "-"  # 0 trials
    assert grid[1][2] == "57.00 ± 7.07"  # "ID" column reads score key "id"


def test_stats_table_grid_ste_spread() -> None:
    grid = ArtifactManager._stats_table_grid(
        ("All",),
        [("hp", [_comp(0.50)["map"], _comp(0.60)["map"]])],
        "ste",
    )

    assert grid[1][1] == "55.00 ± 5.00"  # ste = std / sqrt(n): 7.07 / sqrt(2)


def test_stats_table_grid_single_acc_column() -> None:
    grid = ArtifactManager._stats_table_grid(
        ("I2T",),
        [("hp", [_comp(0.50)["acc"], _comp(0.60)["acc"]])],
        "std",
    )

    assert grid == [["Setting", "I2T"], ["hp (2)", "61.00 ± 7.07"]]


def test_update_stats_tables_writes_pngs(tmp_path, monkeypatch) -> None:
    # "sw" is planned in campaign_metadata.json but has no completed trials in this dataset, so it
    # gets no row (exclusion asserted in test_update_stats_tables_ordered_localized_per_metric);
    # trials are counted by their written final-eval metrics, same as update_metric_stats.
    # bold_high=True + heatmap='fixed' exercise the real matplotlib styling paths (winner bold +
    # heatmap shading) end-to-end.
    dataset = "cub"
    dpath_final = tmp_path / "settings" / "hp" / dataset / "42" / "evals" / "final"
    dpath_final.mkdir(parents=True)
    (dpath_final / "metrics.json").write_text(json.dumps({
        "scores": {"closed_set": {"standard": _scores_grp(_comp(0.50))}},
    }))
    (tmp_path / "campaign_metadata.json").write_text(json.dumps({"settings": ["hp", "sw"], "datasets": [dataset]}))

    monkeypatch.setattr(ArtifactManager, "dpath_campaign", tmp_path)
    monkeypatch.setattr(ArtifactManager, "dataset", dataset)

    ArtifactManager.update_stats_tables("closed_standard", "std", True, False, "fixed", False)

    assert (tmp_path / "stats" / dataset / "map.png").exists()
    assert (tmp_path / "stats" / dataset / "acc.png").exists()


def test_update_stats_tables_ordered_localized_per_metric(tmp_path, monkeypatch) -> None:
    # ordered=True: each png orders setting rows by its own metric's means over ITS dataset's trials
    # only. The bryo data makes cub's local orders the opposite of the cross-dataset ones: cub-local
    # mAP gives a=60 > b=40 -> [a, b] (cross-dataset means 35 vs 40 would say [b, a]), and cub-local
    # acc gives b=80 > a=20 -> [b, a] (cross-dataset means 55 vs 45 would say [a, b]). "c" completed
    # only in bryo -> no row in cub's pngs (blank rows are xlsx-only).
    comp_vals = {  # (setting, dataset) -> (map "all", acc "i2t")
        ("a", "cub"): (0.60, "0.20"), ("a", "bryo"): (0.10, "0.90"),
        ("b", "cub"): (0.40, "0.80"), ("b", "bryo"): (0.40, "0.10"),
        ("c", "bryo"): (0.90, "0.90"),
    }
    for (setting, dataset), (all_v, acc_v) in comp_vals.items():
        dpath_final = tmp_path / "settings" / setting / dataset / "42" / "evals" / "final"
        dpath_final.mkdir(parents=True)
        (dpath_final / "metrics.json").write_text(json.dumps({
            "scores": {"closed_set": {"standard": _scores_grp(_full_comp(all_v, acc_v))}},
        }))
    (tmp_path / "campaign_metadata.json").write_text(
        json.dumps({"settings": ["a", "b", "c"], "datasets": ["cub", "bryo"]})
    )

    grids = []
    monkeypatch.setattr(ArtifactManager, "dpath_campaign", tmp_path)
    monkeypatch.setattr(ArtifactManager, "dataset", "cub")
    monkeypatch.setattr(
        ArtifactManager, "_render_stats_table",
        lambda grid, title, fpath, bold_high, heatmap: grids.append(grid),
    )

    ArtifactManager.update_stats_tables("closed_standard", "std", False, True, None, False)

    grid_map, grid_acc = grids
    assert grid_map[0] == ["Setting", "All", "ID", "OOD", "I2T", "I2I", "T2I"]
    assert [r[0] for r in grid_map[1:]] == ["a (1)", "b (1)"]  # cub-local mAP order; no "c" row
    assert grid_map[1][1] == "60.00"
    assert grid_map[2][1] == "40.00"
    assert grid_acc[0] == ["Setting", "I2T"]
    assert grid_acc[1] == ["b (1)", "80.00"]  # cub-local acc order [b, a]
    assert grid_acc[2] == ["a (1)", "20.00"]
    assert len(grid_acc) == 3  # no "c" row here either


def test_update_metrics_xlsx_writes_stacked_tables(tmp_path, monkeypatch) -> None:
    # two datasets -> two stacked tables sharing the same setting rows (in campaign_metadata order).
    # "hp" has 2 cub trials (mean ± spread) and none in bryo -> a blank "-" row in the Bryozoa table;
    # "sw" has no completed trials anywhere -> no rows at all until its first trial completes.
    settings = ["hp", "sw"]
    for seed, base in (("42", 0.50), ("43", 0.60)):
        dpath_final = tmp_path / "settings" / "hp" / "cub" / seed / "evals" / "final"
        dpath_final.mkdir(parents=True)
        (dpath_final / "metrics.json").write_text(json.dumps({
            "scores": {"closed_set": {"standard": _scores_grp(_comp(base))}},
        }))
    (tmp_path / "campaign_metadata.json").write_text(
        json.dumps({"settings": settings, "datasets": ["cub", "bryo"]})
    )

    monkeypatch.setattr(ArtifactManager, "dpath_campaign", tmp_path)

    ArtifactManager.update_metrics_xlsx("closed_standard", "std", False, False, None, False, False)

    fpath_xlsx = tmp_path / "stats" / "metrics.xlsx"
    assert fpath_xlsx.exists()
    wb = load_workbook(fpath_xlsx)
    ws = wb.active

    def rows_as_lists():
        return [[c.value for c in row] for row in ws.iter_rows()]

    grid = rows_as_lists()
    # campaign banner ("<parent-dir> - <campaign>") + blank row, then the dataset tables, then the
    # always-on cross-dataset mean table (one row per setting) at the bottom. "sw" (no completed
    # trials anywhere) gets no rows; "hp" completed only in cub, so the Bryozoa table still gets its
    # blank "-" row. mean cells are point values (no spread), unlike the per-dataset "± spread".
    assert grid[0][0] == f"{paths['root'].parent.name} - {tmp_path.name}"
    assert grid[1][:7] == [None] * 7  # blank row below the campaign banner
    assert grid[2][0] == "CUB"
    assert grid[3][:7] == ["Setting", "All", "ID", "OOD", "I2T", "I2I", "T2I"]
    assert grid[4][:2] == ["hp (2)", "55.00 ± 7.07"]  # hp: 2 trials
    assert grid[5][:7] == [None] * 7  # spacer row -- no "sw" row
    assert grid[6][0] == "Bryozoa"
    assert grid[7][:7] == ["Setting", "All", "ID", "OOD", "I2T", "I2I", "T2I"]
    assert grid[8][:2] == ["hp (0)", "-"]  # hp's blank entries still added for the trial-less dataset
    assert grid[10][0] == "Mean"  # cross-dataset mean table sits at the bottom
    assert grid[11][0] == "Setting"
    assert grid[12][:7] == ["hp", "55.00", "57.00", "56.00", "58.00", "59.00", "60.00"]
    assert not any(v in ("sw", "sw (0)") for r in grid for v in r)
    # per-seed blocks to the right, one blank separator column apart; "seed <seed>" labels sit in the
    # campaign-banner row; seed blocks have no Mean table and their dataset tables sit in the same
    # rows as the aggregate block's (dataset tables lead everywhere), with plain setting labels (no
    # counts), that seed's raw values, and "-" where the seed's trial hasn't completed
    assert grid[0][8] == "seed 42"
    assert grid[0][16] == "seed 43"
    assert grid[2][8] == "CUB"
    assert grid[3][8:15] == ["Setting", "All", "ID", "OOD", "I2T", "I2I", "T2I"]
    assert grid[4][8:15] == ["hp", "50.00", "52.00", "51.00", "53.00", "54.00", "55.00"]   # seed 42 CUB, aligned with aggregate CUB
    assert grid[6][8] == "Bryozoa"
    assert grid[8][8:15] == ["hp", "-", "-", "-", "-", "-", "-"]                           # seed 42 Bryozoa
    assert grid[4][16:23] == ["hp", "60.00", "62.00", "61.00", "63.00", "64.00", "65.00"]  # seed 43 CUB
    assert grid[10][8] is None  # seed blocks have no Mean table (bottom band stays blank)
    assert not any(r[8] == "Mean" for r in grid)  # no trial-mean table in seed blocks
    assert all(r[7] is None and r[15] is None for r in grid)  # separator columns stay empty
    # snug column widths: longest header/data cell + 2; empty separator columns get a small fixed width
    assert ws.column_dimensions["A"].width == len("Setting") + 2
    assert ws.column_dimensions["B"].width == len("55.00 ± 7.07") + 2
    assert ws.column_dimensions["H"].width == 3
    # bold_high=False: data cells stay unbolded (only header row + setting column bold)
    assert ws.cell(row=5, column=2).font.bold is not True
    # heatmap=None: data cells are left unshaded
    assert ws.cell(row=5, column=2).fill.patternType is None
    # "All Borders": thin black gridlines on every table cell, incl. all cells of the merged title banner
    assert ws.cell(row=3, column=1).border.top.style == "thin"
    assert ws.cell(row=3, column=1).border.top.color.rgb[-6:] == "000000"
    assert ws.cell(row=3, column=7).border.right.color.rgb[-6:] == "000000"  # banner's far merged edge
    assert ws.cell(row=5, column=2).border.left.color.rgb[-6:] == "000000"  # data cell
    # campaign + table titles are left-aligned in their cells
    assert ws.cell(row=1, column=1).alignment.horizontal == "left"
    assert ws.cell(row=3, column=1).alignment.horizontal == "left"
    # 2nd sheet: the accuracy analog (single I2T column per table), same layout/row order.
    # hp's cub trials have acc i2t 56.00/66.00 -> mean 61.00 (± 7.07 in the per-dataset table).
    assert wb.sheetnames == ["Composite mAP", "Composite I2T Accuracy"]
    agrid = [[c.value for c in row] for row in wb["Composite I2T Accuracy"].iter_rows()]
    assert agrid[0][0] == f"{paths['root'].parent.name} - {tmp_path.name}"
    assert agrid[2][0] == "CUB"
    assert agrid[3][:2] == ["Setting", "I2T"]
    assert agrid[4][:2] == ["hp (2)", "61.00 ± 7.07"]
    assert agrid[6][0] == "Bryozoa"
    assert agrid[8][:2] == ["hp (0)", "-"]
    assert agrid[10][0] == "Mean"
    assert agrid[12][:2] == ["hp", "61.00"]
    # acc seed blocks (2-wide, so at cols D-E and G-H)
    assert agrid[0][3] == "seed 42"
    assert agrid[0][6] == "seed 43"
    assert agrid[2][3] == "CUB"
    assert agrid[4][3:5] == ["hp", "56.00"]   # seed 42 CUB, aligned with aggregate CUB
    assert agrid[4][6:8] == ["hp", "66.00"]   # seed 43 CUB
    assert agrid[8][3:5] == ["hp", "-"]       # seed 42 Bryozoa


def test_update_metrics_xlsx_bold_high(tmp_path, monkeypatch) -> None:
    # bold_high=True: the highest-mean setting cell in each score column is bolded. "hp" (base 0.60)
    # outranks "sw" (base 0.50) in every column, so hp's cells bold and sw's do not; "iw" completed
    # only in bryo, so its CUB row is blank "-" -- ignored and never bolded.
    settings = ["hp", "sw", "iw"]
    for setting, dataset, base in (("hp", "cub", 0.60), ("sw", "cub", 0.50), ("iw", "bryo", 0.10)):
        dpath_final = tmp_path / "settings" / setting / dataset / "42" / "evals" / "final"
        dpath_final.mkdir(parents=True)
        (dpath_final / "metrics.json").write_text(json.dumps({
            "scores": {"closed_set": {"standard": _scores_grp(_comp(base))}},
        }))
    (tmp_path / "campaign_metadata.json").write_text(
        json.dumps({"settings": settings, "datasets": ["cub", "bryo"]})
    )

    monkeypatch.setattr(ArtifactManager, "dpath_campaign", tmp_path)

    ArtifactManager.update_metrics_xlsx("closed_standard", "std", True, False, None, False, False)

    ws = load_workbook(tmp_path / "stats" / "metrics.xlsx").active
    # campaign banner + blank row, then the CUB table first: banner row 3, header row 4, setting rows
    # 5/6/7 = hp/sw/iw; score cols B..G = All/ID/OOD/I2T/I2I/T2I (the Mean table sits at the bottom)
    assert ws.cell(row=3, column=1).value == "CUB"
    assert ws.cell(row=4, column=1).value == "Setting"
    for score_col in range(2, 8):
        assert ws.cell(row=5, column=score_col).font.bold is True       # hp wins -> bold
        assert ws.cell(row=6, column=score_col).font.bold is not True   # sw loses -> not bold
        assert ws.cell(row=7, column=score_col).value == "-"            # iw: no cub trials
        assert ws.cell(row=7, column=score_col).font.bold is not True   # "-" never bolds


def test_update_metrics_xlsx_selects_eval_group(tmp_path, monkeypatch) -> None:
    # table_eval_group routes which comp map is read: 'closed_macro' -> scores.closed_set.per_class,
    # not the (different) scores.closed_set.standard values.
    dpath_final = tmp_path / "settings" / "hp" / "cub" / "42" / "evals" / "final"
    dpath_final.mkdir(parents=True)
    (dpath_final / "metrics.json").write_text(json.dumps({
        "scores": {"closed_set": {
            "standard": _scores_grp(_comp(0.50)),   # All -> 50.00
            "per_class": _scores_grp(_comp(0.30)),  # All -> 30.00
        }},
    }))
    (tmp_path / "campaign_metadata.json").write_text(
        json.dumps({"settings": ["hp"], "datasets": ["cub"]})
    )

    monkeypatch.setattr(ArtifactManager, "dpath_campaign", tmp_path)

    ArtifactManager.update_metrics_xlsx("closed_macro", "std", False, False, None, False, False)

    ws = load_workbook(tmp_path / "stats" / "metrics.xlsx").active
    assert ws.cell(row=4, column=2).value == "All"
    assert ws.cell(row=5, column=1).value == "hp (1)"
    assert ws.cell(row=5, column=2).value == "30.00"  # per_class, not standard's 50.00


def _full_comp(all_v: float, acc_v: str = "0.10") -> dict:
    # comp with controllable map "all" + acc "i2t"; other leaves fixed (irrelevant to ordering / heatmap rows)
    return {
        "acc": {"i2t": acc_v},
        "map": {"all": f"{all_v:.4f}", "id": "0.10", "ood": "0.10", "i2t": "0.10", "i2i": "0.10", "t2i": "0.10"},
    }


def test_update_metrics_xlsx_ordered_per_sheet_metric(tmp_path, monkeypatch) -> None:
    # ordered=True: each sheet orders its columns by its own metric's Mean-table first row, so the
    # two sheets may disagree. Campaign order is [a, b]; mAP mean-All gives a=mean(20,40)=30.00 <
    # b=mean(40,40)=40.00 -> mAP sheet flips to [b, a], while acc mean-I2T gives a=80.00 > b=20.00
    # -> accuracy sheet keeps [a, b].
    for setting, acc_v, cub_all, bryo_all in (("a", "0.80", 0.20, 0.40), ("b", "0.20", 0.40, 0.40)):
        for dataset, all_v in (("cub", cub_all), ("bryo", bryo_all)):
            dpath_final = tmp_path / "settings" / setting / dataset / "42" / "evals" / "final"
            dpath_final.mkdir(parents=True)
            (dpath_final / "metrics.json").write_text(json.dumps({
                "scores": {"closed_set": {"standard": _scores_grp(_full_comp(all_v, acc_v))}},
            }))
    (tmp_path / "campaign_metadata.json").write_text(
        json.dumps({"settings": ["a", "b"], "datasets": ["cub", "bryo"]})
    )

    monkeypatch.setattr(ArtifactManager, "dpath_campaign", tmp_path)

    ArtifactManager.update_metrics_xlsx("closed_standard", "std", False, True, None, False, False)

    wb = load_workbook(tmp_path / "stats" / "metrics.xlsx")
    grid = [[c.value for c in r] for r in wb.active.iter_rows()]
    # campaign banner + blank row, then the dataset tables (Mean at the bottom); setting rows ordered
    # by the mean table's "All" column -> b before a
    assert grid[2][0] == "CUB"
    assert grid[3][:7] == ["Setting", "All", "ID", "OOD", "I2T", "I2I", "T2I"]
    assert grid[4][:2] == ["b (1)", "40.00"]
    assert grid[5][:2] == ["a (1)", "20.00"]
    assert grid[7][0] == "Bryozoa"
    assert grid[9][:2] == ["b (1)", "40.00"]
    assert grid[10][:2] == ["a (1)", "40.00"]
    assert grid[12][0] == "Mean"
    assert grid[14][:2] == ["b", "40.00"]
    assert grid[15][:2] == ["a", "30.00"]
    # the seed block's rows are pinned to the sheet's mean-derived order too (CUB table, aligned rows)
    assert grid[0][8] == "seed 42"
    assert [grid[4][8], grid[5][8]] == ["b", "a"]
    # the accuracy sheet orders by its own acc mean-'I2T' column -> [a, b], unlike the mAP sheet
    agrid = [[c.value for c in r] for r in wb["Composite I2T Accuracy"].iter_rows()]
    assert agrid[3][:2] == ["Setting", "I2T"]
    assert agrid[4][:2] == ["a (1)", "80.00"]
    assert agrid[5][:2] == ["b (1)", "20.00"]
    assert agrid[12][0] == "Mean"
    assert agrid[14][:2] == ["a", "80.00"]
    assert agrid[15][:2] == ["b", "20.00"]
    assert agrid[4][3:5] == ["a", "80.00"]  # acc seed block keeps the acc sheet's [a, b] order (aligned rows)


def _fill_rgb(ws, row, col):
    # last 6 hex chars (RGB) of a cell's fill, or None when the cell is unshaded
    fill = ws.cell(row=row, column=col).fill
    return None if fill.patternType is None else fill.fgColor.rgb[-6:]


def test_update_metrics_xlsx_heatmap_scaled(tmp_path, monkeypatch) -> None:
    # heatmap='scaled': each column's min -> white (#ffffff), max -> #ff5533, rest linearly
    # interpolated. One dataset, so the mean "All" column mirrors the values 20/50/80; "d" (no
    # completed trials anywhere) gets no row at all.
    for setting, all_v in (("a", 0.20), ("b", 0.50), ("c", 0.80)):
        dpath_final = tmp_path / "settings" / setting / "cub" / "42" / "evals" / "final"
        dpath_final.mkdir(parents=True)
        (dpath_final / "metrics.json").write_text(json.dumps({
            "scores": {"closed_set": {"standard": _scores_grp(_full_comp(all_v))}},
        }))
    (tmp_path / "campaign_metadata.json").write_text(
        json.dumps({"settings": ["a", "b", "c", "d"], "datasets": ["cub"]})  # "d" has no trials
    )

    monkeypatch.setattr(ArtifactManager, "dpath_campaign", tmp_path)

    ArtifactManager.update_metrics_xlsx("closed_standard", "std", False, False, "scaled", False, False)

    ws = load_workbook(tmp_path / "stats" / "metrics.xlsx").active
    # campaign banner + blank row; CUB table first: banner row 3, header row 4, "All" column is col B,
    # setting rows 5/6/7 = a/b/c (values 20/50/80)
    assert ws.cell(row=8, column=1).value is None  # spacer right after c -> no "d" row
    assert _fill_rgb(ws, 5, 2) == "FFFFFF"  # column min -> white
    assert _fill_rgb(ws, 6, 2) == "FFAA99"  # midpoint (t=0.5) -> interpolated
    assert _fill_rgb(ws, 7, 2) == "FF5533"  # column max -> #ff5533
    # the trailing Mean table is shaded too (setting rows 11/12/13)
    assert _fill_rgb(ws, 11, 2) == "FFFFFF"
    assert _fill_rgb(ws, 13, 2) == "FF5533"


def test_update_metrics_xlsx_heatmap_fixed(tmp_path, monkeypatch) -> None:
    # heatmap='fixed': value/100 maps to white->#ff5533 regardless of the column's other cells;
    # 20/50/80 -> #ffddd6 / #ffaa99 / #ff775c.
    for setting, all_v in (("a", 0.20), ("b", 0.50), ("c", 0.80)):
        dpath_final = tmp_path / "settings" / setting / "cub" / "42" / "evals" / "final"
        dpath_final.mkdir(parents=True)
        (dpath_final / "metrics.json").write_text(json.dumps({
            "scores": {"closed_set": {"standard": _scores_grp(_full_comp(all_v))}},
        }))
    (tmp_path / "campaign_metadata.json").write_text(
        json.dumps({"settings": ["a", "b", "c"], "datasets": ["cub"]})
    )

    monkeypatch.setattr(ArtifactManager, "dpath_campaign", tmp_path)

    ArtifactManager.update_metrics_xlsx("closed_standard", "std", False, False, "fixed", False, False)

    ws = load_workbook(tmp_path / "stats" / "metrics.xlsx").active
    # campaign banner + blank row; CUB table first: banner row 3, header row 4, "All" column is col B,
    # setting rows 5/6/7 = a/b/c
    assert _fill_rgb(ws, 5, 2) == "FFDDD6"  # 20 -> t=0.20
    assert _fill_rgb(ws, 6, 2) == "FFAA99"  # 50 -> t=0.50
    assert _fill_rgb(ws, 7, 2) == "FF775C"  # 80 -> t=0.80


def _prim_scores_grp() -> dict:
    # _scores_grp with distinct per-partition primitive values (comp base 0.50)
    grp = _scores_grp(_comp(0.50))
    grp["id"] = {"map": {"i2t": "0.61", "i2i": "0.62", "t2i": "0.63"}, "acc": {"i2t": "0.64"}}
    grp["ood"] = {"map": {"i2t": "0.71", "i2i": "0.72", "t2i": "0.73"}, "acc": {"i2t": "0.74"}}
    return grp


_PRIM_MAP_HEADER = ["Setting", "All", "ID", "OOD", "I2T", "I2I", "T2I",
                    "ID I2T", "ID I2I", "ID T2I", "OOD I2T", "OOD I2I", "OOD T2I"]


def test_update_metrics_xlsx_prim_scores(tmp_path, monkeypatch) -> None:
    # prim_scores=True appends the per-partition primitive score columns: ID/OOD x I2T/I2I/T2I on the
    # mAP sheet, ID I2T / OOD I2T on the accuracy sheet -- in every table, incl. the seed blocks
    dpath_final = tmp_path / "settings" / "hp" / "cub" / "42" / "evals" / "final"
    dpath_final.mkdir(parents=True)
    (dpath_final / "metrics.json").write_text(json.dumps({
        "scores": {"closed_set": {"standard": _prim_scores_grp()}},
    }))
    (tmp_path / "campaign_metadata.json").write_text(json.dumps({"settings": ["hp"], "datasets": ["cub"]}))

    monkeypatch.setattr(ArtifactManager, "dpath_campaign", tmp_path)

    ArtifactManager.update_metrics_xlsx("closed_standard", "std", False, False, None, True, False)

    wb = load_workbook(tmp_path / "stats" / "metrics.xlsx")
    ws = wb.active
    grid = [[c.value for c in r] for r in ws.iter_rows()]
    # mAP banners split: unmerged title + grey merged 'Composite Scores'/'Primitive Scores' group headers
    assert grid[2][:2] == ["CUB", "Composite Scores"]
    assert grid[2][7] == "Primitive Scores"
    assert grid[6][:2] == ["Mean", "Composite Scores"]
    assert grid[6][7] == "Primitive Scores"
    assert "B3:G3" in {str(m) for m in ws.merged_cells.ranges} and "H3:M3" in {str(m) for m in ws.merged_cells.ranges}
    assert ws.cell(row=3, column=2).fill.fgColor.rgb[-6:] == "EAEAEA"  # group headers get the header grey
    assert ws.cell(row=3, column=1).fill.patternType is None           # title cell stays unfilled
    assert grid[3][:13] == _PRIM_MAP_HEADER
    assert grid[4][:13] == ["hp (1)", "50.00", "52.00", "51.00", "53.00", "54.00", "55.00",
                            "61.00", "62.00", "63.00", "71.00", "72.00", "73.00"]  # CUB row
    assert grid[8][7:13] == ["61.00", "62.00", "63.00", "71.00", "72.00", "73.00"]  # Mean row (bottom)
    # seed block starts after the 13-wide aggregate + separator; its CUB table row-aligns with the aggregate's
    assert grid[0][14] == "seed 42"
    assert grid[2][14:16] == ["CUB", "Composite Scores"]
    assert grid[2][21] == "Primitive Scores"
    assert grid[3][14:27] == _PRIM_MAP_HEADER
    assert grid[4][21:27] == ["61.00", "62.00", "63.00", "71.00", "72.00", "73.00"]
    # the accuracy sheet keeps full-width merged title banners (no group headers)
    ws_acc = wb["Composite I2T Accuracy"]
    agrid = [[c.value for c in r] for r in ws_acc.iter_rows()]
    assert agrid[2][:2] == ["CUB", None]
    assert "A3:D3" in {str(m) for m in ws_acc.merged_cells.ranges}
    assert agrid[3][:4] == ["Setting", "I2T", "ID I2T", "OOD I2T"]
    assert agrid[4][:4] == ["hp (1)", "56.00", "64.00", "74.00"]  # CUB row
    assert agrid[8][:4] == ["hp", "56.00", "64.00", "74.00"]  # Mean row (bottom)
    assert agrid[0][5] == "seed 42"


def test_update_stats_tables_prim_scores(tmp_path, monkeypatch) -> None:
    # prim_scores=True appends the per-partition primitive score columns to the png grids too
    dpath_final = tmp_path / "settings" / "hp" / "cub" / "42" / "evals" / "final"
    dpath_final.mkdir(parents=True)
    (dpath_final / "metrics.json").write_text(json.dumps({
        "scores": {"closed_set": {"standard": _prim_scores_grp()}},
    }))
    (tmp_path / "campaign_metadata.json").write_text(json.dumps({"settings": ["hp"], "datasets": ["cub"]}))

    grids = []
    monkeypatch.setattr(ArtifactManager, "dpath_campaign", tmp_path)
    monkeypatch.setattr(ArtifactManager, "dataset", "cub")
    monkeypatch.setattr(
        ArtifactManager, "_render_stats_table",
        lambda grid, title, fpath, bold_high, heatmap: grids.append(grid),
    )

    ArtifactManager.update_stats_tables("closed_standard", "std", False, False, None, True)

    grid_map, grid_acc = grids
    assert grid_map[0] == _PRIM_MAP_HEADER
    assert grid_map[1] == ["hp (1)", "50.00", "52.00", "51.00", "53.00", "54.00", "55.00",
                           "61.00", "62.00", "63.00", "71.00", "72.00", "73.00"]
    assert grid_acc[0] == ["Setting", "I2T", "ID I2T", "OOD I2T"]
    assert grid_acc[1] == ["hp (1)", "56.00", "64.00", "74.00"]


def test_update_metrics_xlsx_baseline_overrides(tmp_path, monkeypatch) -> None:
    # baseline_overrides=True renders a "Baseline Overrides" config table in a left column band that
    # the score blocks shift right past, aligned with the aggregate block's bottom Mean table (whose
    # Setting column labels its rows): one column per overridden param (union of the settings'
    # overrides.json keys, first-seen order), values resolved from each setting's
    # setting_metadata.json -- "-" when the param is absent there (inert under that config: sw has
    # loss2.mix 0.0, so clean_metadata dropped its loss2 subtree). Config cells get no
    # winner-bold/heatmap styling despite bold_high/heatmap on.
    for setting, base, overrides, meta in (
        ("hp", 0.50, {"loss2.mix": 0.3, "loss2.targ": "phylo"},
         {"loss": {"targ": "sw"}, "loss2": {"mix": 0.3, "targ": "phylo"}}),
        ("sw", 0.40, {"loss.targ": "sw"},
         {"loss": {"targ": "sw"}}),
    ):
        dpath_setting = tmp_path / "settings" / setting
        dpath_final = dpath_setting / "cub" / "42" / "evals" / "final"
        dpath_final.mkdir(parents=True)
        (dpath_final / "metrics.json").write_text(json.dumps({
            "scores": {"closed_set": {"standard": _scores_grp(_comp(base))}},
        }))
        (dpath_setting / "overrides.json").write_text(json.dumps(overrides))
        (dpath_setting / "setting_metadata.json").write_text(json.dumps(meta))
    (tmp_path / "campaign_metadata.json").write_text(json.dumps({"settings": ["hp", "sw"], "datasets": ["cub"]}))

    monkeypatch.setattr(ArtifactManager, "dpath_campaign", tmp_path)

    ArtifactManager.update_metrics_xlsx("closed_standard", "std", True, False, "fixed", False, True)

    wb = load_workbook(tmp_path / "stats" / "metrics.xlsx")
    ws = wb.active
    grid = [[c.value for c in r] for r in ws.iter_rows()]
    # score blocks shift right past the 3-wide overrides band + separator col D: aggregate at E..K
    assert grid[0][0] == f"{paths['root'].parent.name} - {tmp_path.name}"  # campaign banner stays top-left
    assert grid[2][4] == "CUB"
    assert grid[7][4] == "Mean"
    # overrides table in the left band, aligned with the Mean table; param cols in first-seen order
    # (hp's overrides, then sw's); no Setting column of its own
    assert grid[7][0] == "Baseline Overrides"
    assert "A8:C8" in {str(m) for m in ws.merged_cells.ranges}
    assert grid[8][:3] == ["loss2.mix", "loss2.targ", "loss.targ"]
    assert grid[8][4:6] == ["Setting", "All"]  # Mean header shares the row
    assert grid[9][:3] == ["0.3", "phylo", "sw"]
    assert grid[9][4] == "hp"  # labeled by the Mean's Setting column
    assert grid[10][:3] == ["-", "-", "sw"]  # loss2.* inert for sw -> "-"
    assert grid[10][4] == "sw"
    # param-name header styled like other headers; config cells skip score styling entirely
    assert ws.cell(row=9, column=1).fill.fgColor.rgb[-6:] == "EAEAEA"
    assert ws.cell(row=9, column=1).font.bold is True
    assert ws.cell(row=10, column=1).fill.patternType is None
    assert ws.cell(row=10, column=1).font.bold is not True
    assert all(r[3] is None for r in grid)  # separator col D stays empty
    # seed block one separator past the aggregate score block (cols E..K)
    assert grid[0][12] == "seed 42"
    assert grid[2][12] == "CUB"
    # same treatment on the accuracy sheet (2-wide score tables at E..F, seed at col I)
    ws_acc = wb["Composite I2T Accuracy"]
    agrid = [[c.value for c in r] for r in ws_acc.iter_rows()]
    assert agrid[2][4] == "CUB"
    assert agrid[7][4] == "Mean"
    assert agrid[7][0] == "Baseline Overrides"
    assert "A8:C8" in {str(m) for m in ws_acc.merged_cells.ranges}
    assert agrid[8][:3] == ["loss2.mix", "loss2.targ", "loss.targ"]
    assert agrid[9][:3] == ["0.3", "phylo", "sw"]
    assert agrid[10][:3] == ["-", "-", "sw"]
    assert agrid[0][7] == "seed 42"


def test_load_base_eval_cache_misses_when_entry_lacks_needed_pieces(tmp_path, monkeypatch) -> None:
    # entries carry only what the caching trial computed -- a trial must read an entry missing a
    # piece it needs (projections for viz, embs for pooled) as a miss (recompute + upgrade the
    # entry) rather than hit the missing piece downstream in _write_base_eval; leaner trials
    # still reuse the entry
    fpath = tmp_path / "combo.pkl"
    monkeypatch.setattr(ArtifactManager, "base_eval_cache_fpath", lambda cfg: fpath)
    entry = {"metrics": {"scores": {"comp": {"map": {"all": "0.50"}}}}, "projections": None, "embs": None}

    assert ArtifactManager.load_base_eval_cache(None, require_projections=False, require_embs=False) is None  # no file for this combo

    save_pickle(entry, fpath)
    assert ArtifactManager.load_base_eval_cache(None, require_projections=True, require_embs=False) is None  # metrics-only entry, viz trial
    assert ArtifactManager.load_base_eval_cache(None, require_projections=False, require_embs=False) == entry

    entry_viz = {**entry, "projections": {"pca_id": [0.0]}}
    save_pickle(entry_viz, fpath)
    assert ArtifactManager.load_base_eval_cache(None, require_projections=True, require_embs=False) == entry_viz
    assert ArtifactManager.load_base_eval_cache(None, require_projections=True, require_embs=True) is None  # no embs, pooled trial

    entry_pooled = {**entry_viz, "embs": {"embs_id": [0.0]}}
    save_pickle(entry_pooled, fpath)
    assert ArtifactManager.load_base_eval_cache(None, require_projections=True, require_embs=True) == entry_pooled


def test_save_base_eval_cache_writes_per_combo_file(tmp_path, monkeypatch) -> None:
    # each save ingests the npz files compute_projections wrote into this trial's evals/_base/
    # (absent for non-viz trials -> None) and writes its combo's entry to that combo's own file,
    # leaving other combos' files untouched
    dpath_cache = tmp_path / "base_eval_cache"
    monkeypatch.setattr(ArtifactManager, "base_eval_cache_fpath", lambda cfg: dpath_cache / "combo.pkl")
    monkeypatch.setattr(ArtifactManager, "dpath_trial", tmp_path / "trial")
    dpath_base = tmp_path / "trial" / "evals" / "_base"
    dpath_base.mkdir(parents=True)
    np.savez(dpath_base / "projections.npz", pca_id=np.arange(3))
    eval_metrics = {"scores": {"comp": {"map": {"all": 0.5}}}, "loss_raw": {"id": 0.7, "ood": None}}

    ArtifactManager.save_base_eval_cache(None, eval_metrics)

    entry = load_pickle(dpath_cache / "combo.pkl")
    assert entry["metrics"] == {"scores": {"comp": {"map": {"all": "0.5000"}}}}  # loss_raw stripped
    assert list(entry["projections"]) == ["pca_id"]
    assert entry["embs"] is None

    monkeypatch.setattr(ArtifactManager, "base_eval_cache_fpath", lambda cfg: dpath_cache / "combo2.pkl")
    monkeypatch.setattr(ArtifactManager, "dpath_trial", tmp_path / "trial2")  # no _base npzs -> non-viz trial

    ArtifactManager.save_base_eval_cache(None, eval_metrics)

    assert sorted(p.name for p in dpath_cache.iterdir()) == ["combo.pkl", "combo2.pkl"]
    assert load_pickle(dpath_cache / "combo2.pkl")["projections"] is None


def test_base_eval_key_normalizes_family_inert_components() -> None:
    # non_causal is CLIP-only, vis_proj is SigLIP-only, and seed only enters through the random
    # init of a linear/mlp vis_proj head -- inert components read as None so equivalent configs
    # share one cache entry
    def cfg(model_type, non_causal=False, vis_proj=None):
        return SimpleNamespace(
            arch={"model_type": model_type, "clip": {"non_causal": non_causal}, "siglip": {"vis_proj": vis_proj}},
            img_norm="default", dataset="cub", split="dev",
            text_template={"train": "train", "eval": "sci"}, seed=42,
        )

    assert ArtifactManager.base_eval_key(cfg("siglip_vitb16")) == \
        ("siglip_vitb16", "default", "cub", "dev", None, "sci", None, None)  # headless: seed shared
    assert ArtifactManager.base_eval_key(cfg("siglip_vitb16", vis_proj="mlp")) == \
        ("siglip_vitb16", "default", "cub", "dev", None, "sci", "mlp", 42)  # random head: seed kept
    assert ArtifactManager.base_eval_key(cfg("clip_vitb16", non_causal=True)) == \
        ("clip_vitb16", "default", "cub", "dev", True, "sci", None, None)
    # non_causal true vs false are two separate cached readings
    assert ArtifactManager.base_eval_key(cfg("clip_vitb16", non_causal=True)) != \
        ArtifactManager.base_eval_key(cfg("clip_vitb16", non_causal=False))
    # the combo key serializes to the flat per-combo cache filename
    fpath = ArtifactManager.base_eval_cache_fpath(cfg("siglip_vitb16", vis_proj="mlp"))
    assert fpath.parent.name == "base_eval_cache"
    assert fpath.name == "siglip_vitb16__default__cub__dev__None__sci__mlp__42.pkl"


def test_format_and_merge_mem_running_max() -> None:
    # bytes -> 'used/total GB' (GiB), and merge keeps the higher-used reading per key -- a running
    # max across snapshots; None (no reading yet) is always superseded
    snap = format_mem({"ram": (4.2 * 2**30, 128 * 2**30), "vram": (37.5 * 2**30, 79.3 * 2**30)})
    assert snap == {"ram": "4.2/128.0 GB", "vram": "37.5/79.3 GB"}

    assert merge_mem({"ram": None, "vram": None}, snap) == snap

    later = {"ram": "6.0/128.0 GB", "vram": "12.0/79.3 GB"}
    assert merge_mem(snap, later) == {"ram": "6.0/128.0 GB", "vram": "37.5/79.3 GB"}
