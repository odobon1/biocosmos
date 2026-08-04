"""
Campaign reporting/presentation: metric-stats aggregation, composite-score summary tables
(map.png / acc.png), metrics.xlsx, and per-trial learning-curve plots. Everything here renders
from artifacts already on disk and reads its paths from ArtifactManager; trial/checkpoint state
I/O lives in utils/train.py.
"""

import matplotlib
matplotlib.use("Agg")
import matplotlib.pyplot as plt
import matplotlib.gridspec as gridspec
from matplotlib.lines import Line2D
from matplotlib.ticker import FuncFormatter, FormatStrFormatter
import numpy as np
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from utils.ddp import rank0
from utils.train import ArtifactManager
from utils.utils import (
    paths,
    save_json,
    save_json_listview,
    load_json,
    DATASET_ALIAS2NAME,
)

import pdb


# table_eval_group -> (scores set_key, group key, human-readable name); drives map.png and metrics.xlsx
_TABLE_EVAL_GROUPS = {
    "closed_standard": ("closed_set", "standard", "Closed-Set, Standard"),
    "closed_macro": ("closed_set", "per_class", "Closed-Set, Macro"),
    "full_standard": ("full_set", "standard", "Full-Set, Standard"),
    "full_macro": ("full_set", "per_class", "Full-Set, Macro"),
}

# "Hardware Performance" table columns; _HW_LABELS key the per-trial dicts built by _collect_hw,
# _HW_CRASH_LABELS the per-setting crash totals it reads from setting_metadata.json
_HW_LABELS = ("Time Trial", "Mean Time Train", "Mean Time Eval", "Peak RAM", "Peak VRAM")
_HW_CRASH_LABELS = ("Total Crashes RAM", "Total Crashes VRAM", "Total Crashes Other")


def _spread(nums, spread_type):
    spread = nums.std(ddof=1)
    return spread / np.sqrt(len(nums)) if spread_type == "ste" else spread

def _aggregate_metric_stats(values, spread_type, percent=True):
    first = values[0]
    if first is None:  # leaf is None for every trial (e.g. loss_raw["ood"] is never computed)
        return None
    if isinstance(first, dict):
        return {
            k: _aggregate_metric_stats(
                [v[k] for v in values], spread_type, percent=percent and k not in ("loss_raw", "sim", "targ")
            )
            for k in first
        }
    if len(values) == 1:
        return values[0]
    nums = np.array([float(v) for v in values])
    mean = nums.mean()
    spread = _spread(nums, spread_type)
    if percent:
        return f"{mean * 100:.2f} ± {spread * 100:.2f}"
    return f"{mean:.4f} ± {spread:.4f}"

def _listview_metric_stats(values, percent=True):
    first = values[0]
    if first is None:  # leaf is None for every trial (mirrors _aggregate_metric_stats)
        return None
    if isinstance(first, dict):
        return {
            k: _listview_metric_stats(
                [v[k] for v in values], percent=percent and k not in ("loss_raw", "sim", "targ")
            )
            for k in first
        }
    if percent:
        return [f"{float(v) * 100:.2f}" for v in values]
    return [f"{float(v):.4f}" for v in values]

@rank0
def update_metric_stats(spread_type):
    dpath_dataset = ArtifactManager.dpath_setting / ArtifactManager.dataset
    metric_dicts = []
    for dpath_trial in sorted(dpath_dataset.iterdir()):
        # a written final-eval metrics file is the signal a trial finished; the `complete` flag is
        # marked later (campaign_runner, after a clean exit) so it can't gate this aggregation
        fpath_metrics = dpath_trial / "evals/final/metrics.json"
        if not fpath_metrics.exists():
            continue
        metrics = load_json(fpath_metrics)
        metrics.pop("n_samps_seen", None)
        metric_dicts.append(metrics)

    if not metric_dicts:
        return

    n_trials = len(metric_dicts)
    stats = {
        "n_trials": n_trials,
        **_aggregate_metric_stats(metric_dicts, spread_type),
    }
    listview = {
        "n_trials": n_trials,
        **_listview_metric_stats(metric_dicts),
    }
    dpath_stats = dpath_dataset / "stats"
    dpath_stats.mkdir(parents=True, exist_ok=True)
    save_json(stats, dpath_stats / "metrics.json")
    save_json_listview(listview, dpath_stats / "metrics_listview.json")

def _stats_table_grid(labels, setting_score_maps, spread_type):
    """Build a composite-score table's cell grid from [(setting, [score dict per completed
    trial]), ...]: a header row of 'Setting' + one column per label in `labels`, then one row
    per setting -- '<setting> (n_trials)' + each label's cell (read from the score dicts by
    its lowercased key): '-' (0 trials), 'XX.XX' (1 trial, mean), or 'XX.XX ± XX.XX'
    (>1 trial, mean ± spread)."""
    grid = [["Setting", *labels]]
    for setting, score_maps in setting_score_maps:
        row = [f"{setting} ({len(score_maps)})"]
        for label in labels:
            nums = np.array([float(score_map[label.lower()]) for score_map in score_maps]) * 100
            if len(nums) == 0:
                row.append("-")
            elif len(nums) == 1:
                row.append(f"{nums[0]:.2f}")
            else:
                row.append(f"{nums.mean():.2f} ± {_spread(nums, spread_type):.2f}")
        grid.append(row)
    return grid

def _collect_comps(settings, datasets, set_key, grp):
    """Each (setting, dataset)'s completed-trial score maps, keyed by trial seed (the trial dir
    name; empty dict -> no trials yet): per trial a {'map': ..., 'acc': ...} pair of flat
    label->score dicts merging the comp scores with the per-partition primitives ('id i2t' ...
    'ood t2i'), so table labels map to keys by lowercasing. A written final-eval metrics file is
    the completion signal, same as update_metric_stats."""
    comps_by = {}
    for setting in settings:
        for dataset in datasets:
            dpath_dataset = ArtifactManager.dpath_campaign / "settings" / setting / dataset
            comps = {}
            if dpath_dataset.exists():
                for dpath_trial in sorted(dpath_dataset.iterdir()):
                    fpath_metrics = dpath_trial / "evals/final/metrics.json"
                    if fpath_metrics.exists():
                        scores_grp = load_json(fpath_metrics)["scores"][set_key][grp]
                        comps[dpath_trial.name] = {
                            "map": {**scores_grp["comp"]["map"],
                                    **{f"{p} {m}": scores_grp[p]["map"][m] for p in ("id", "ood") for m in ("i2t", "i2i", "t2i")}},
                            "acc": {**scores_grp["comp"]["acc"],
                                    **{f"{p} i2t": scores_grp[p]["acc"]["i2t"] for p in ("id", "ood")}},
                        }
            comps_by[(setting, dataset)] = comps
    return comps_by

def _collect_hw(settings, datasets):
    """Each (setting, dataset)'s completed-trial hardware/wall-clock readings, parsed from
    trial_metadata.json and keyed by trial seed (the trial dir name, like _collect_comps): one
    {_HW_LABELS label -> float} dict per trial -- runtime.trial / runtime.train.mean /
    runtime.eval.mean are float-seconds strings, memory.ram / memory.vram are 'used/total GB'
    strings (numerator taken). A written final-eval metrics file is the completion signal, same
    as _collect_comps. Also each setting's n_crashes totals ({'ram'/'vram'/'other' -> int},
    summed across all its trials -- seeds + datasets, completed or not) from
    setting_metadata.json, whose counters survive the trial-dir wipes that reset
    trial_metadata's."""
    hw_by = {}
    for setting in settings:
        for dataset in datasets:
            dpath_dataset = ArtifactManager.dpath_campaign / "settings" / setting / dataset
            trials = {}
            if dpath_dataset.exists():
                for dpath_trial in sorted(dpath_dataset.iterdir()):
                    if (dpath_trial / "evals/final/metrics.json").exists():
                        meta = load_json(dpath_trial / "trial_metadata.json")
                        trials[dpath_trial.name] = {
                            "Time Trial": float(meta["runtime"]["trial"]),
                            "Mean Time Train": float(meta["runtime"]["train"]["mean"]),
                            "Mean Time Eval": float(meta["runtime"]["eval"]["mean"]),
                            "Peak RAM": float(meta["memory"]["ram"].split("/")[0]),
                            "Peak VRAM": float(meta["memory"]["vram"].split("/")[0]),
                        }
            hw_by[(setting, dataset)] = trials
    crashes_by = {
        setting: load_json(ArtifactManager.dpath_campaign / "settings" / setting / "setting_metadata.json")["n_crashes"]
        for setting in settings
    }
    return hw_by, crashes_by

def _score_labels(prim_scores):
    """(mAP labels, acc labels) for the stats tables; prim_scores appends the per-partition
    primitive score columns (ID/OOD x modality) to each."""
    map_labels = ("All", "ID", "OOD", "I2T", "I2I", "T2I")
    acc_labels = ("I2T",)
    if prim_scores:
        map_labels += ("ID I2T", "ID I2I", "ID T2I", "OOD I2T", "OOD I2I", "OOD T2I")
        acc_labels += ("ID I2T", "OOD I2T")
    return map_labels, acc_labels

def _cross_dataset_means(settings, datasets, comps_by, score_key, labels):
    """xmeans[(setting, label)]: arithmetic mean, across datasets with completed trials, of that
    setting/label's per-dataset mean comp score (percent), read from comp[score_key][label.lower()]
    (score_key: 'map' or 'acc'); None when no dataset has trials."""

    def dataset_means(setting, label):
        key = label.lower()
        return [np.mean([float(comp[score_key][key]) for comp in comps_by[(setting, dataset)].values()]) * 100
                for dataset in datasets if comps_by[(setting, dataset)]]

    xmeans = {}
    for setting in settings:
        for label in labels:
            vals = dataset_means(setting, label)
            xmeans[(setting, label)] = np.mean(vals) if vals else None
    return xmeans

def _order_settings(settings, xmeans, label):
    # order setting rows by the mean for `label`, descending (ties keep campaign order); callers
    # filter out settings with no completed trials before ordering, so every mean is numeric
    return sorted(settings, key=lambda s: xmeans[(s, label)], reverse=True)

def _col_styles(grid, bold_high):
    """Per-column data-cell styling for one rendered table, shared by the png and xlsx tables:
    styles[c] for each score-label column c -- row -> mean for numeric cells ('-' skipped), the
    bold-winner rows (highest mean, ties included; empty unless bold_high), and the column's
    min/max mean for scaled heatmaps."""
    styles = {}
    for c in range(1, len(grid[0])):
        means = {r: float(grid[r][c].split(" ± ")[0]) for r in range(1, len(grid)) if grid[r][c] != "-"}
        winners = set()
        if bold_high and means:
            top = max(means.values())
            winners = {r for r, m in means.items() if m == top}
        col_min = min(means.values()) if means else 0.0
        col_max = max(means.values()) if means else 0.0
        styles[c] = (means, winners, col_min, col_max)
    return styles

def _heat_hex(heatmap, mean, col_min, col_max):
    """Heatmap cell color as 'RRGGBB': linear white (#ffffff) -> #ff5533 interpolation -- 'fixed'
    maps a fixed 0.00 -> 100.00, 'scaled' maps the column's min -> max (a single-value or
    all-equal column -> lowest color)."""
    if heatmap == "fixed":
        frac = mean / 100.0
    elif col_max > col_min:  # scaled across the column's data cells
        frac = (mean - col_min) / (col_max - col_min)
    else:  # scaled but column has one value (or all equal) -> lowest color
        frac = 0.0
    frac = max(0.0, min(1.0, frac))
    g = round(255 - (255 - 0x55) * frac)
    b = round(255 - (255 - 0x33) * frac)
    return f"FF{g:02X}{b:02X}"

def _render_stats_table(grid, title, fpath, bold_high, heatmap):
    fig, ax = plt.subplots(figsize=(1.2 + 1.5 * (len(grid[0]) - 1), 0.7 + 0.3 * len(grid)))
    ax.axis("off")
    ax.set_title(title, fontsize=11, pad=12)
    table = ax.table(cellText=grid, loc="center", cellLoc="center")
    table.auto_set_font_size(False)
    table.set_fontsize(9)
    table.scale(1, 1.5)
    table.auto_set_column_width(list(range(len(grid[0]))))
    styles = _col_styles(grid, bold_high)
    for (row, col), cell in table.get_celld().items():
        if row == 0 or col == 0:
            cell.set_text_props(fontweight="bold")
            cell.set_facecolor("#eaeaea")
            continue
        means, winners, col_min, col_max = styles[col]
        if row in winners:
            cell.set_text_props(fontweight="bold")
        if heatmap and row in means:
            cell.set_facecolor(f"#{_heat_hex(heatmap, means[row], col_min, col_max)}")
    fig.savefig(fpath, dpi=200, bbox_inches="tight")
    plt.close(fig)

@rank0
def update_stats_tables(table_eval_group, spread_type, bold_high, ordered, heatmap, prim_scores):
    """Render the campaign-level composite-score summary tables for this trial's dataset to
    artifacts/<campaign>/stats/<dataset>/map.png (comp mAP: All/ID/OOD/I2T/I2I/T2I score
    columns) and acc.png (comp I2T accuracy: single I2T column) -- prim_scores appends the
    per-partition primitive score columns (ID/OOD x modality) to both: one row per setting with >= 1
    completed trial in this dataset (settings without local trials are omitted -- no blank rows
    in the pngs), stats aggregated across each setting's completed trials. bold_high/ordered/
    heatmap style the
    tables the same way as metrics.xlsx: bold_high bolds each score column's highest-mean cell
    (ties included; '-' cells ignored), ordered orders each table's setting rows by its own
    metric's mean over THIS dataset's completed trials (map.png by the mAP 'All' column,
    acc.png by the acc 'I2T' column) -- localized per dataset, independent of the cross-dataset
    order used in metrics.xlsx -- heatmap shades score cells white->#ff5533 (None/scaled/fixed
    as in update_metrics_xlsx). Re-rendered at each trial completion."""
    set_key, grp, group_name = _TABLE_EVAL_GROUPS[table_eval_group]

    settings = load_json(ArtifactManager.dpath_campaign / "campaign_metadata.json")["settings"]
    dataset = ArtifactManager.dataset
    comps_by = _collect_comps(settings, (dataset,), set_key, grp)
    # a setting gets a row only once it has >= 1 completed trial in THIS dataset (no blank rows)
    settings = [s for s in settings if comps_by[(s, dataset)]]

    def ordered_settings(score_key, label):
        # localized order: this dataset's per-setting trial means (single-dataset degenerate
        # case of _cross_dataset_means), not the cross-dataset means backing the xlsx order
        means = _cross_dataset_means(settings, (dataset,), comps_by, score_key, (label,))
        return _order_settings(settings, means, label)

    settings_map = ordered_settings("map", "All") if ordered else settings
    settings_acc = ordered_settings("acc", "I2T") if ordered else settings
    map_labels, acc_labels = _score_labels(prim_scores)

    title_suffix = f" -- {DATASET_ALIAS2NAME[dataset]} ({group_name})"
    dpath_stats = ArtifactManager.dpath_campaign / "stats" / dataset
    dpath_stats.mkdir(parents=True, exist_ok=True)

    grid_map = _stats_table_grid(
        map_labels,
        [(setting, [comp["map"] for comp in comps_by[(setting, dataset)].values()]) for setting in settings_map],
        spread_type,
    )
    _render_stats_table(grid_map, f"Composite mAP{title_suffix}", dpath_stats / "map.png", bold_high, heatmap)

    grid_acc = _stats_table_grid(
        acc_labels,
        [(setting, [comp["acc"] for comp in comps_by[(setting, dataset)].values()]) for setting in settings_acc],
        spread_type,
    )
    _render_stats_table(grid_acc, f"Composite I2T Accuracy{title_suffix}", dpath_stats / "acc.png", bold_high, heatmap)

@rank0
def update_metrics_xlsx(table_eval_group, spread_type, bold_high, ordered, heatmap, prim_scores, baseline_overrides, hw_perf):
    """Write artifacts/<campaign>/stats/metrics.xlsx: two sheets, 'Composite mAP' (comp map scores,
    All/ID/OOD/I2T/I2I/T2I score columns) and 'Composite I2T Accuracy' (comp acc, single I2T
    column) -- prim_scores appends the per-partition primitive score columns (ID/OOD x modality)
    to both sheets' tables, and splits each mAP-sheet table banner into the title cell (first
    column, unmerged) plus grey merged 'Composite Scores' / 'Primitive Scores' group headers over
    their column groups (the accuracy sheet keeps full-width merged title banners). Each sheet
    opens with a bold '<repo-parent-dir> - <campaign>' title cell (e.g.
    'bc_dev - dev') and a blank row, then stacks one table per campaign dataset vertically -- a bold left-aligned
    title banner, then a table of header row 'Setting' + one column per score label and one
    '<setting> (n_trials)' row per setting, then a blank spacer row before the next dataset -- with
    the always-shown 'Mean' summary table at the bottom: one row per setting, each cell the
    arithmetic mean, across datasets with completed trials, of that setting/label's per-dataset
    mean (a point value, no spread). Cells are '-' (0 trials), 'XX.XX' (1 trial, mean) or 'XX.XX ± XX.XX'
    (>1 trial, mean ± spread), aggregated across each setting's completed trials'
    scores.<set_key>.<grp>.comp for the eval group selected by table_eval_group. A setting gets
    rows only once it has >= 1 completed trial in some dataset -- it then appears in every table of
    both sheets, with blank '-' rows in dataset tables lacking its trials; settings with no
    completed trials anywhere are omitted entirely. When bold_high is
    True, the highest-mean setting cell in each score column is bolded (ties included; '-' cells
    ignored). When ordered is True, each sheet's setting rows are ordered by its own metric's
    Mean-table first score column -- 'All' for mAP, 'I2T' for accuracy -- descending, settings with
    no completed trials anywhere last; when False, rows keep the fixed campaign_metadata order.
    Within a sheet one row order is shared across all tables, but the two sheets' orders may differ.
    heatmap shades each score cell white->#ff5533 by intensity: None leaves cells unshaded; 'scaled'
    maps each score column's min->max to white->#ff5533; 'fixed' maps a fixed 0.00->100.00 to
    white->#ff5533. '-' cells are never shaded. To the right of this aggregate block sit per-seed
    blocks (one blank separator column apart): a 'seed <seed>' label in the campaign-banner row,
    then the per-dataset tables only (no Mean summary) built from that seed's trials alone,
    sitting in the same rows as the aggregate block's dataset tables -- plain setting labels (no
    trial counts), single-trial 'XX.XX' cells, '-' where that seed's trial hasn't completed --
    sharing the aggregate block's setting rows/order. baseline_overrides adds a 'Baseline
    Overrides' config table to each sheet, in a left column band that the score blocks shift
    right past (one blank separator column between), vertically aligned with the aggregate
    block's bottom Mean table so the Mean's Setting column labels its rows: one column per param
    overridden in the campaign's baseline_overrides (union of the settings' overrides.json keys,
    first-seen order), each cell the setting's effective value resolved from its
    config.json -- '-' when the param is absent there, the signal that it is inert
    under that configuration (e.g. loss2.* with loss2.mix 0.0). Params whose effective value is
    identical across every setting row are omitted (they differentiate nothing); when every param
    is uniform the table is omitted entirely. This table gets no
    winner-bold/heatmap styling. hw_perf adds a companion 'Hardware Performance' table to the
    right of every scores table on the mAP sheet only (one blank separator column between the
    two; the accuracy sheet gets none), row-aligned with its scores table so the scores Setting
    column labels its rows, and likewise unstyled: Time Trial / Mean Time Train / Mean Time Eval
    (whole seconds) and Peak RAM / Peak VRAM (whole GB) columns, per-trial readings parsed from
    trial_metadata.json (float-seconds runtime strings; 'used/total GB' memory strings, numerator
    taken), every cell rounded to the nearest int. Each companion aggregates the same trials as
    its scores table: dataset-table companions the mean across that dataset's completed trials
    ('-' rows where the setting has none there), seed-block companions that seed's single-trial
    readings ('-' where its trial hasn't completed), and the Mean-table companion the mean across
    datasets with completed trials of the setting's per-dataset trial means -- plus Total Crashes
    RAM / VRAM / Other columns (Mean companion only, since they don't decompose per dataset/seed),
    each cell the setting's crash total of that cause across all its trials (seeds + datasets,
    completed or not), read from setting_metadata.json's n_crashes. Column widths hug each
    column's longest header/data cell (banner/label text overflows); blank separator columns get
    a small ~square width. Regenerated at each trial completion."""
    set_key, grp, _ = _TABLE_EVAL_GROUPS[table_eval_group]
    metadata = load_json(ArtifactManager.dpath_campaign / "campaign_metadata.json")
    settings, datasets = metadata["settings"], metadata["datasets"]

    comps_by = _collect_comps(settings, datasets, set_key, grp)
    # a setting gets rows only once it has >= 1 completed trial in some dataset; it then appears in
    # every dataset table (blank '-' row where that dataset has no trials for it yet)
    settings = [s for s in settings if any(comps_by[(s, dataset)] for dataset in datasets)]
    seeds = sorted({seed for comps in comps_by.values() for seed in comps}, key=int)

    if baseline_overrides:
        dpath_settings = ArtifactManager.dpath_campaign / "settings"
        okeys = []  # union of overridden params, first-seen order across settings (campaign order)
        for s in settings:
            for key in load_json(dpath_settings / s / "overrides.json"):
                if key not in okeys:
                    okeys.append(key)
        metadata_by = {s: load_json(dpath_settings / s / "config.json") for s in settings}

        def override_value(setting, key):
            # a param absent from the setting's metadata is inert under that configuration -> '-'
            node = metadata_by[setting]
            for part in key.split("."):
                if not isinstance(node, dict) or part not in node:
                    return "-"
                node = node[part]
            return str(node)

        # a param whose effective value is identical across every setting row differentiates
        # nothing -- drop the column (and with it the whole table when no column survives)
        okeys = [key for key in okeys if len({override_value(s, key) for s in settings}) > 1]

    hw_by, crashes_by = _collect_hw(settings, datasets) if hw_perf else (None, None)

    def build_blocks(score_key, labels, with_hw):
        """The sheet's blocks, left to right: (label, [(title, cell grid, hw grid), ...]) -- the
        aggregate block (label None): one table per campaign dataset, then the always-shown 'Mean'
        cross-dataset summary table at the bottom; then one block per completed seed (label
        'seed <seed>'): the per-dataset tables only (no Mean summary), built from that seed's
        trials alone -- plain setting labels (no trial counts), single-trial 'XX.XX' cells, '-'
        where that seed's trial hasn't completed. Setting rows are shared across all blocks --
        when ordered, pinned to the aggregate Mean-table's first score column (labels[0]),
        descending. Each table's hw grid is its 'Hardware Performance' companion (None when
        with_hw is off): _HW_LABELS header + one value row per setting (same row order, no
        Setting column of its own), cells the rounded mean over the same trials as the scores
        table beside it ('-' when the setting has none there); the Mean companion instead holds
        the cross-dataset mean of per-dataset trial means plus the _HW_CRASH_LABELS per-setting
        crash-total columns. Also returns ogrid, the 'Baseline Overrides' grid (param-name
        header + one value row per setting in this sheet's row order), or None when disabled or
        nothing is overridden."""
        xmeans = _cross_dataset_means(settings, datasets, comps_by, score_key, labels)
        rows = _order_settings(settings, xmeans, labels[0]) if ordered else settings

        def hw_grid(readings_by_row):
            # one companion grid: per setting row a list of that row's per-trial readings dicts
            # (empty -> '-' cells), meaned per label and rounded to the nearest int
            grid = [list(_HW_LABELS)]
            for readings in readings_by_row:
                if readings:
                    grid.append([str(round(np.mean([r[label] for r in readings]))) for label in _HW_LABELS])
                else:
                    grid.append(["-"] * len(_HW_LABELS))
            return grid

        tables = []
        for dataset in datasets:
            grid = _stats_table_grid(
                labels,
                [(setting, [comp[score_key] for comp in comps_by[(setting, dataset)].values()]) for setting in rows],
                spread_type,
            )
            hw = hw_grid([list(hw_by[(s, dataset)].values()) for s in rows]) if with_hw else None
            tables.append((DATASET_ALIAS2NAME[dataset], grid, hw))
        xgrid = [["Setting", *labels]]
        for s in rows:
            xgrid.append([s] + [f"{xmeans[(s, label)]:.2f}" for label in labels])
        xhw = None
        if with_hw:
            xhw = [list(_HW_LABELS) + list(_HW_CRASH_LABELS)]
            for s in rows:
                xhw.append([
                    str(round(np.mean([np.mean([trial[label] for trial in hw_by[(s, dataset)].values()])
                                       for dataset in datasets if hw_by[(s, dataset)]])))
                    for label in _HW_LABELS
                ] + [str(crashes_by[s][kind]) for kind in ("ram", "vram", "other")])
        tables.append(("Mean", xgrid, xhw))
        blocks = [(None, tables)]

        for seed in seeds:
            stables = []
            for dataset in datasets:
                grid = [["Setting", *labels]]
                for s in rows:
                    comp = comps_by[(s, dataset)].get(seed)
                    grid.append([s] + ["-" if comp is None else f"{float(comp[score_key][label.lower()]) * 100:.2f}"
                                       for label in labels])
                hw = None
                if with_hw:
                    hw = hw_grid([[hw_by[(s, dataset)][seed]] if seed in hw_by[(s, dataset)] else []
                                  for s in rows])
                stables.append((DATASET_ALIAS2NAME[dataset], grid, hw))
            blocks.append((f"seed {seed}", stables))

        ogrid = None
        if baseline_overrides and okeys:
            # header of param names only -- no setting column; the rows align with (and are
            # labeled by) the aggregate Mean table's setting rows
            ogrid = [list(okeys)]
            for s in rows:
                ogrid.append([override_value(s, key) for key in okeys])
        return blocks, ogrid

    bold = Font(bold=True)
    center = Alignment(horizontal="center", vertical="center")
    left = Alignment(horizontal="left", vertical="center")
    header_fill = PatternFill("solid", fgColor="EAEAEA")
    thin = Side(style="thin", color="000000")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    def write_sheet(ws, blocks, groups, ogrid):
        """groups: None -> each table's banner is its title merged across the full table width;
        else [(group_title, n_group_cols), ...] -> the title sits unmerged in the block's first
        column, followed by one grey merged group-header cell per group (e.g. 'Composite Scores'
        over the composite columns, 'Primitive Scores' over the primitive columns). Each table's
        hw grid (None when off) renders as its 'Hardware Performance' companion, one blank
        separator column to the scores table's right: a merged title banner in the scores banner
        row, then the _HW_LABELS header + value rows row-aligned with the scores rows (the scores
        Setting column labels them); companion cells get no winner-bold/heatmap styling. ogrid
        (None or a header + one-value-row-per-setting grid, sheet row order) renders as the
        'Baseline Overrides' left column band, one blank separator column after it, with the
        score blocks all shifted right past it -- vertically aligned with the aggregate block's
        bottom Mean table so the Mean's Setting column labels its rows; band cells likewise get
        no winner-bold/heatmap styling."""
        n_cols = len(blocks[0][1][0][1][0])  # corner + one col per score label (same for every grid in the sheet)
        widths = {}  # col idx -> longest header/data cell text (banner/label cells overflow instead)

        campaign = ws.cell(row=1, column=1, value=f"{paths['root'].parent.name} - {ArtifactManager.dpath_campaign.name}")
        campaign.font = bold
        campaign.alignment = left

        bands = [("Baseline Overrides", ogrid)] if ogrid else []
        offset = sum(len(g[0]) + 1 for _, g in bands)  # left bands + their separator columns
        col0 = 1 + offset  # blocks side by side, one blank separator column apart
        for block_label, tables in blocks:
            if block_label is not None:
                label_cell = ws.cell(row=1, column=col0, value=block_label)  # campaign-banner row, atop the block
                label_cell.font = bold
                label_cell.alignment = left
            # banner row + blank row above the tables; dataset tables lead in every block, so they
            # sit in the same rows across blocks (only the aggregate has the trailing Mean table)
            row = 3
            for title_text, grid, hw in tables:
                title = ws.cell(row=row, column=col0, value=title_text)
                title.font = bold
                title.alignment = left
                for c in range(col0, col0 + n_cols):  # border every cell of the banner so the merged ranges' edges all render
                    ws.cell(row=row, column=c).border = border
                if groups is None:
                    ws.merge_cells(start_row=row, start_column=col0, end_row=row, end_column=col0 + n_cols - 1)
                else:
                    widths[col0] = max(widths.get(col0, 0), len(title_text))  # unmerged title must fit its column
                    gcol = col0 + 1
                    for group_title, n_group in groups:
                        for c in range(gcol, gcol + n_group):  # fill every cell so the merged range renders grey
                            ws.cell(row=row, column=c).fill = header_fill
                        gcell = ws.cell(row=row, column=gcol, value=group_title)
                        gcell.font = bold
                        gcell.alignment = center
                        ws.merge_cells(start_row=row, start_column=gcol, end_row=row, end_column=gcol + n_group - 1)
                        gcol += n_group
                if hw:
                    hw_col0 = col0 + n_cols + 1  # companion sits one blank separator column right of the scores table
                    hw_title = ws.cell(row=row, column=hw_col0, value="Hardware Performance")
                    hw_title.font = bold
                    hw_title.alignment = left
                    for c in range(hw_col0, hw_col0 + len(hw[0])):
                        ws.cell(row=row, column=c).border = border
                    ws.merge_cells(start_row=row, start_column=hw_col0, end_row=row, end_column=hw_col0 + len(hw[0]) - 1)
                row += 1

                styles = _col_styles(grid, bold_high)
                for r, grid_row in enumerate(grid):
                    for c, val in enumerate(grid_row):
                        cell = ws.cell(row=row, column=col0 + c, value=val)
                        cell.alignment = center
                        cell.border = border
                        widths[col0 + c] = max(widths.get(col0 + c, 0), len(val))
                        if r == 0 or c == 0:
                            cell.font = bold
                            cell.fill = header_fill
                            continue
                        means, winners, col_min, col_max = styles[c]
                        if r in winners:
                            cell.font = bold
                        if heatmap and r in means:
                            cell.fill = PatternFill("solid", fgColor=_heat_hex(heatmap, means[r], col_min, col_max))
                    if hw:  # companion rows align with the scores rows (header + one row per setting)
                        for c, val in enumerate(hw[r]):
                            cell = ws.cell(row=row, column=hw_col0 + c, value=val)
                            cell.alignment = center
                            cell.border = border
                            widths[hw_col0 + c] = max(widths.get(hw_col0 + c, 0), len(val))
                            if r == 0:
                                cell.font = bold
                                cell.fill = header_fill
                    row += 1
                row += 1  # blank spacer row between tables
            # widest table decides the block's width (the Mean companion carries extra crash columns)
            col0 += n_cols + max((len(hw[0]) + 1 for _, _, hw in tables if hw), default=0) + 1

        # banner row of the aggregate block's bottom Mean table (dataset tables precede it)
        band_row = 3 + sum(len(grid) + 2 for _, grid, _ in blocks[0][1][:-1])
        band_col = 1
        for band_title, band_grid in bands:
            b_cols = len(band_grid[0])
            title = ws.cell(row=band_row, column=band_col, value=band_title)
            title.font = bold
            title.alignment = left
            for c in range(band_col, band_col + b_cols):
                ws.cell(row=band_row, column=c).border = border
            ws.merge_cells(start_row=band_row, start_column=band_col, end_row=band_row, end_column=band_col + b_cols - 1)
            row = band_row + 1
            for r, grid_row in enumerate(band_grid):
                for c, val in enumerate(grid_row):
                    cell = ws.cell(row=row, column=band_col + c, value=val)
                    cell.alignment = center
                    cell.border = border
                    widths[band_col + c] = max(widths.get(band_col + c, 0), len(val))
                    if r == 0:  # header row (param names)
                        cell.font = bold
                        cell.fill = header_fill
                row += 1
            band_col += b_cols + 1

        for c in range(1, max(widths) + 1):
            # snug fit to each column's longest cell; blank separator columns get a small ~square width
            ws.column_dimensions[get_column_letter(c)].width = widths[c] + 2 if c in widths else 3

    map_labels, acc_labels = _score_labels(prim_scores)
    # with prim_scores, mAP-sheet banners split into title + 'Composite Scores'/'Primitive Scores'
    # group headers; the accuracy sheet keeps full-width title banners
    map_groups = [("Composite Scores", 6), ("Primitive Scores", 6)] if prim_scores else None
    wb = Workbook()
    ws_map = wb.active
    ws_map.title = "Composite mAP"
    map_blocks, map_ogrid = build_blocks("map", map_labels, hw_perf)
    write_sheet(ws_map, map_blocks, map_groups, map_ogrid)
    # hw companions are mAP-sheet only: the accuracy sheet keeps just the overrides band
    acc_blocks, acc_ogrid = build_blocks("acc", acc_labels, False)
    write_sheet(wb.create_sheet("Composite I2T Accuracy"), acc_blocks, None, acc_ogrid)

    dpath_stats = ArtifactManager.dpath_campaign / "stats"
    dpath_stats.mkdir(parents=True, exist_ok=True)
    wb.save(dpath_stats / "metrics.xlsx")


def _samples_seen_tick_formatter(value, _pos):
    return f"{value / 1_000_000:g}"

@rank0
def plot_metrics(
        data_tracker,
        dpath_trial,
        fontsize_axes=12,
        fontsize_ticks=8,
        fontsize_legend=8,
        subplot_border_width=1,
        figsize=(10, 16),
        height_ratios=[2, 2, 2, 2, 2, 1, 1, 1],
    ):
    data = data_tracker.data
    data_epoch = data["epoch"]
    data_eval = data["eval"]
    title_suffix = f" -- {ArtifactManager.dpath_setting.name}, {DATASET_ALIAS2NAME[ArtifactManager.dataset]}"

    # eval panels (retrieval / n-shot / accuracy) are populated only when eval ran;
    # train panels (loss / grad norm / lr) plot whenever train data is present (e.g. train_pt=trainval).
    partitions = [k for k in data_eval.get("scores", {}).get("closed_set", {}).get("standard", {}).keys() if k != "comp"]

    x_eval = data_eval["n_samps_seen"]
    x_train = data_epoch["n_samps_seen"]

    bucket_partition = "id" if "id" in partitions else None
    bucket_comp_keys_standard = [
        key for key in data_eval.get("scores", {}).get("closed_set", {}).get("standard", {}).get(bucket_partition, {}).get("map", {}).get("n-shot", {}).keys()
    ]
    bucket_comp_keys_full_set = [
        key
        for key in data_eval.get("scores", {}).get("full_set", {}).get("standard", {}).get(bucket_partition, {}).get("map", {}).get("n-shot", {}).keys()
    ]

    plot_composite_metrics(
        data_epoch,
        data_eval,
        x_train,
        x_eval,
        dpath_trial,
        partitions,
        bucket_partition,
        bucket_comp_keys_standard,
        fontsize_axes,
        fontsize_ticks,
        fontsize_legend,
        subplot_border_width,
        figsize,
        height_ratios,
        partition_metric_group="standard",
        full_set=False,
        retrieval_ylabel="mAP Scores",
        accuracy_ylabel="I2T Accuracy",
        nshot_accuracy_ylabel="n-shot Accuracy (ID)",
        plot_title=f"Train Metrics{title_suffix}",
        output_filename="closed_standard.png",
    )

    plot_composite_metrics(
        data_epoch,
        data_eval,
        x_train,
        x_eval,
        dpath_trial,
        partitions,
        bucket_partition,
        bucket_comp_keys_standard,
        fontsize_axes,
        fontsize_ticks,
        fontsize_legend,
        subplot_border_width,
        figsize,
        height_ratios,
        partition_metric_group="per_class",
        full_set=False,
        retrieval_ylabel="Macro mAP Scores",
        accuracy_ylabel="I2T Per-Class Accuracy",
        nshot_accuracy_ylabel="n-shot Per-Class\nAccuracy (ID)",
        plot_title=f"Train Metrics (Macro){title_suffix}",
        output_filename="closed_macro.png",
    )

    plot_composite_metrics(
        data_epoch,
        data_eval,
        x_train,
        x_eval,
        dpath_trial,
        partitions,
        bucket_partition,
        bucket_comp_keys_full_set,
        fontsize_axes,
        fontsize_ticks,
        fontsize_legend,
        subplot_border_width,
        figsize,
        height_ratios,
        partition_metric_group="standard",
        full_set=True,
        retrieval_ylabel="Full-Set mAP Scores",
        accuracy_ylabel="Full-Set I2T Accuracy",
        nshot_accuracy_ylabel="Full-Set n-shot Accuracy (ID)",
        plot_title=f"Train Metrics (Full-Set){title_suffix}",
        output_filename="full_standard.png",
    )

    plot_composite_metrics(
        data_epoch,
        data_eval,
        x_train,
        x_eval,
        dpath_trial,
        partitions,
        bucket_partition,
        bucket_comp_keys_full_set,
        fontsize_axes,
        fontsize_ticks,
        fontsize_legend,
        subplot_border_width,
        figsize,
        height_ratios,
        partition_metric_group="per_class",
        full_set=True,
        retrieval_ylabel="Full-Set Macro mAP Scores",
        accuracy_ylabel="Full-Set I2T Per-Class Accuracy",
        nshot_accuracy_ylabel="Full-Set n-shot Per-Class\nAccuracy (ID)",
        plot_title=f"Train Metrics (Macro Full-Set){title_suffix}",
        output_filename="full_macro.png",
    )

def plot_composite_metrics(
    data_epoch,
    data_eval,
    x_train,
    x_eval,
    dpath_trial,
    partitions,
    bucket_partition,
    bucket_comp_keys,
    fontsize_axes,
    fontsize_ticks,
    fontsize_legend,
    subplot_border_width,
    figsize,
    height_ratios,
    partition_metric_group,
    full_set,
    retrieval_ylabel,
    accuracy_ylabel,
    nshot_accuracy_ylabel,
    plot_title,
    output_filename,
):
    fig = plt.figure(figsize=figsize)
    gs = gridspec.GridSpec(len(height_ratios), 1, height_ratios=height_ratios, hspace=0)

    ax0 = fig.add_subplot(gs[0, 0])

    id_partition = "id" if "id" in partitions else None
    ood_partition = "ood" if "ood" in partitions else None
    retrieval_specs = (
        ("i2t", "I2T", "blue"),
        ("i2i", "I2I", "red"),
        ("t2i", "T2I", "green"),
    )
    set_key = "full_set" if full_set else "closed_set"
    style_specs = (
        (id_partition, "ID", "-"),
        (ood_partition, "OOD", "--"),
    )
    for partition, partition_label, linestyle in style_specs:
        if partition is None:
            continue
        partition_group_scores = data_eval.get("scores", {}).get(set_key, {}).get(partition_metric_group, {}).get(partition, {})
        partition_group_scores = partition_group_scores.get("map", {})
        for metric_name, metric_label, color in retrieval_specs:
            if metric_name in partition_group_scores:
                ax0.plot(
                    x_eval,
                    partition_group_scores[metric_name],
                    label=f"{partition_label} {metric_label}",
                    color=color,
                    linestyle=linestyle,
                )

    ax0.set_ylabel(retrieval_ylabel, fontsize=fontsize_axes, fontweight="bold")
    ax0.set_ylim(0, 1)
    if partitions:
        ax0.legend(loc="lower right", fontsize=fontsize_legend)
    ax0.grid(True)
    ax0.tick_params(labelbottom=False, labelsize=fontsize_ticks)

    ax1 = fig.add_subplot(gs[1, 0], sharex=ax0)
    is_macro_plot = partition_metric_group == "per_class"
    id_mode_scores = data_eval.get("scores", {}).get(set_key, {}).get(partition_metric_group, {}).get(bucket_partition, {})
    nshot_key = "n-shot"
    if is_macro_plot:
        nshot_ylabel = "Full-Set n-shot Macro mAP (ID)" if full_set else "n-shot Macro mAP (ID)"
    else:
        nshot_ylabel = "Full-Set n-shot mAP (ID)" if full_set else "n-shot mAP (ID)"
    comp_nshot = id_mode_scores.get("map", {}).get(nshot_key, {})
    if bucket_comp_keys:
        for key in reversed(bucket_comp_keys):
            label = key
            maybe_plot(ax1, x_eval, comp_nshot, key, label, linestyle=":")
        if comp_nshot:
            ax1.legend(loc="lower right", fontsize=fontsize_legend)
    ax1.set_ylabel(nshot_ylabel, fontsize=fontsize_axes, fontweight="bold")
    ax1.set_ylim(0, 1)
    ax1.grid(True)
    ax1.tick_params(labelbottom=False, labelsize=fontsize_ticks)

    ax2 = fig.add_subplot(gs[2, 0], sharex=ax0)
    for partition in partitions:
        partition_group_scores = data_eval.get("scores", {}).get(set_key, {}).get(partition_metric_group, {}).get(partition, {})
        partition_group_scores = partition_group_scores.get("acc", {})
        if "i2t" in partition_group_scores:
            ax2.plot(
                x_eval,
                partition_group_scores["i2t"],
                label="-".join([s.upper() if i == 0 else s.title() for i, s in enumerate(partition.split("_"))]),
            )
    ax2.set_ylabel(accuracy_ylabel, fontsize=fontsize_axes, fontweight="bold")
    ax2.set_ylim(0, 1)
    if partitions:
        ax2.legend(loc="lower right", fontsize=fontsize_legend)
    ax2.grid(True)
    ax2.tick_params(labelbottom=False, labelsize=fontsize_ticks)

    ax3 = fig.add_subplot(gs[3, 0], sharex=ax0)
    comp_nshot_acc = id_mode_scores.get("acc", {}).get("n-shot", {})
    if bucket_comp_keys:
        for key in reversed(bucket_comp_keys):
            maybe_plot(ax3, x_eval, comp_nshot_acc, key, key, linestyle=":")
        if comp_nshot_acc:
            ax3.legend(loc="lower right", fontsize=fontsize_legend)
    ax3.set_ylabel(nshot_accuracy_ylabel, fontsize=fontsize_axes, fontweight="bold")
    ax3.set_ylim(0, 1)
    ax3.grid(True)
    ax3.tick_params(labelbottom=False, labelsize=fontsize_ticks)

    ax4 = fig.add_subplot(gs[4, 0], sharex=ax0)
    if len(data_epoch.get("loss_train", [])) == len(x_train):
        ax4.plot(x_train, data_epoch["loss_train"], label="Train Loss")
    if len(data_epoch.get("loss_raw_train", [])) == len(x_train):
        ax4.plot(x_train, data_epoch["loss_raw_train"], label="Train Loss (Raw)")
    for partition in partitions:
        if len(data_eval.get("loss_raw", {}).get(partition, [])) == len(x_eval):
            ax4.plot(
                x_eval,
                data_eval["loss_raw"][partition],
                label=f'{"-".join([s.upper() if i == 0 else s.title() for i, s in enumerate(partition.split("_"))])} Val Loss',
            )
    ax4.set_ylabel("Loss", fontsize=fontsize_axes, fontweight="bold")
    ax4.set_yscale("log")
    ax4.minorticks_on()
    ax4.grid(which="minor", axis="y")
    ax4.legend(loc="upper right", fontsize=fontsize_legend)
    ax4.grid(True)
    ax4.tick_params(labelbottom=False, labelsize=fontsize_ticks)

    ax5 = fig.add_subplot(gs[5, 0], sharex=ax0)
    if len(data_epoch.get("grad_norm_model", [])) == len(x_train):
        ax5.plot(x_train, data_epoch["grad_norm_model"], color="green")
    ax5.set_ylabel("Grad Norm", fontsize=fontsize_axes, fontweight="bold")
    ax5.set_yscale("log")
    ax5.minorticks_on()
    ax5.grid(which="minor", axis="y")
    ax5.grid(True)
    ax5.tick_params(labelbottom=False, labelsize=fontsize_ticks)

    ax6 = fig.add_subplot(gs[6, 0], sharex=ax0)
    color_sim = "tab:blue"
    color_targ = "tab:orange"
    # targets first, then similarity; min/max solid, mean dashed, median dotted
    for stat_prefix, stat_color in (("targ", color_targ), ("sim", color_sim)):
        for stat_name, stat_linestyle in (("min", "-"), ("max", "-"), ("mean", "--"), ("median", ":")):
            stat_key = f"{stat_prefix}_{stat_name}"
            if len(data_epoch.get(stat_key, [])) == len(x_train):
                ax6.plot(x_train, data_epoch[stat_key], color=stat_color, linestyle=stat_linestyle, linewidth=1.0)
    ax6.set_ylabel("Similarity / Target", fontsize=fontsize_axes, fontweight="bold")
    ax6.set_ylim(-1.0, 1.0)
    ax6.legend(
        handles=[
            Line2D([0], [0], color=color_sim, lw=1.0, label="Similarity"),
            Line2D([0], [0], color=color_targ, lw=1.0, label="Target"),
            Line2D([0], [0], color="gray", lw=1.0, linestyle="-", label="Min/Max"),
            Line2D([0], [0], color="gray", lw=1.0, linestyle="--", label="Mean"),
            Line2D([0], [0], color="gray", lw=1.0, linestyle=":", label="Median"),
        ],
        loc="upper center",
        ncol=5,
        fontsize=fontsize_legend,
    )
    ax6.grid(True)
    ax6.tick_params(labelbottom=False, labelsize=fontsize_ticks)

    ax7 = fig.add_subplot(gs[7, 0], sharex=ax0)
    if len(data_epoch.get("lr", [])) == len(x_train):
        ax7.plot(x_train, data_epoch["lr"])
    ax7.set_ylabel("Learning Rate", fontsize=fontsize_axes, fontweight="bold")
    ax7.ticklabel_format(style="sci", axis="y", scilimits=(0, 0))
    ax7.yaxis.set_offset_position("right")
    ax7.yaxis.set_major_formatter(FormatStrFormatter("%.1e"))
    ax7.yaxis.get_offset_text().set_visible(False)
    ax7.set_xlabel("Samples Seen (M)", fontsize=fontsize_axes, fontweight="bold")
    ax7.xaxis.set_major_formatter(FuncFormatter(_samples_seen_tick_formatter))
    ax7.grid(True)
    ax7.tick_params(labelsize=fontsize_ticks)

    for ax in (ax0, ax1, ax2, ax3, ax4, ax5, ax6, ax7):
        ax.label_outer()

    for idx_ax, ax in enumerate((ax0, ax1, ax2, ax3, ax4, ax5, ax6, ax7)):
        for spine in ax.spines.values():
            spine.set_linewidth(subplot_border_width)
            spine.set_edgecolor("black")
        if idx_ax % 2 == 1:
            ax.yaxis.set_label_position("right")
            ax.yaxis.tick_right()

    fig.suptitle(plot_title, fontweight="bold", y=0.98, fontsize=20)
    plt.subplots_adjust(hspace=0)
    plt.tight_layout()
    plots_dir = dpath_trial / "learning_curves"
    plots_dir.mkdir(parents=True, exist_ok=True)
    fig.savefig(plots_dir / output_filename, dpi=300)
    plt.close(fig)

def maybe_plot(ax, x, data, key, label, **kwargs):
    """
    Helper for plot_metrics() (N-Shot Composites)
    """
    if key in data and len(data[key]) > 0:
        ax.plot(x, data[key], label=label, **kwargs)
