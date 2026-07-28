import matplotlib
matplotlib.use("Agg")
import matplotlib.pyplot as plt
import matplotlib.gridspec as gridspec
from matplotlib.lines import Line2D
from matplotlib.ticker import FuncFormatter, FormatStrFormatter
from dataclasses import asdict
from datetime import datetime, timezone
import os
import time
import random
import numpy as np
import torch
import shutil
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from utils.utils import (
    paths,
    save_pickle,
    load_pickle,
    save_json,
    save_json_listview,
    load_json,
    TimeTracker,
    Timer,
    DATASET_ALIAS2NAME,
)
from utils.ddp import rank0

import pdb


def format_scores(scores):
    if scores is None:
        return None
    if isinstance(scores, dict):
        return {k: format_scores(v) for k, v in scores.items()}
    return f"{float(scores):.4f}"

def parse_scores(scores):
    if scores is None:
        return None
    if isinstance(scores, dict):
        return {k: parse_scores(v) for k, v in scores.items()}
    return float(scores)

def format_mem(mem):
    # {"ram": (used_bytes, total_bytes), "vram": ...} -> {"ram": "4.2/128.0 GB", "vram": ...}
    return {k: f"{used / 2**30:.1f}/{total / 2**30:.1f} GB" for k, (used, total) in mem.items()}

def merge_mem(mem_prev, mem_new):
    """Per key keep whichever formatted 'used/total GB' reading has the higher used -- a running max
    across snapshots (kept as a pair so used/total stay from the same reading if totals ever differ
    across relaunches). None (no reading yet) is always superseded."""
    def used(s):
        return float(s.split("/")[0])
    return {
        k: mem_new[k] if mem_prev[k] is None or used(mem_new[k]) > used(mem_prev[k]) else mem_prev[k]
        for k in mem_new
    }


class TrialData:

    def __init__(self, dpath_trial):

        self.fpath_data = dpath_trial / "data_trial.pkl"

        self.data_epoch = {
            "n_samps_seen": [],
            "lr": [],
            "loss_train": [],
            "loss_raw_train": [],
            "grad_norm_model": [],
            "sim_min": [],
            "sim_max": [],
            "sim_median": [],
            "sim_mean": [],
            "targ_min": [],
            "targ_max": [],
            "targ_median": [],
            "targ_mean": [],
        }
        self.data_eval = {
            "n_samps_seen": [],
        }
        self.data = {
            "epoch": self.data_epoch,
            "eval": self.data_eval,
        }

        self.n_evals = 0

        self.eval_metrics = None  # most recent eval metrics
        self.time_eval = None  # most recent eval time

        self.timer_trial = Timer()
        self.timer_trial.start()

    def update_train_batch(self, n_samps_seen, lr=None, loss_train=None, loss_raw_train=None, grad_norm_model=None, batch_stats=None):

        self.data_epoch["n_samps_seen"].append(n_samps_seen)

        if lr is not None:
            self.data_epoch["lr"].append(lr)
        if loss_train is not None:
            self.data_epoch["loss_train"].append(loss_train)
        if loss_raw_train is not None:
            self.data_epoch["loss_raw_train"].append(loss_raw_train)
        if grad_norm_model is not None:
            self.data_epoch["grad_norm_model"].append(grad_norm_model)
        if batch_stats is not None:
            for stat_key, stat_value in batch_stats.items():
                self.data_epoch[stat_key].append(stat_value)

    def update_eval(self, n_samps_seen):

        def append_nested(dst, src):
            for score_name, score_value in src.items():
                if isinstance(score_value, dict):
                    if score_name not in dst:
                        dst[score_name] = {}
                    append_nested(dst[score_name], score_value)
                else:
                    if score_name not in dst:
                        dst[score_name] = []
                    dst[score_name].append(score_value)

        self.n_evals += 1

        self.data_eval["n_samps_seen"].append(n_samps_seen)

        for k, v in self.eval_metrics.items():
            if k not in self.data_eval:
                self.data_eval[k] = {}
            append_nested(self.data_eval[k], v)

    @classmethod
    def resume(cls, dpath_trial, trial_state):
        obj = cls(dpath_trial)
        obj.data = load_pickle(obj.fpath_data)
        obj.data_epoch = obj.data["epoch"]
        obj.data_eval = obj.data["eval"]
        obj.n_evals = trial_state["n_evals"]
        obj.timer_trial = Timer()
        obj.timer_trial.set_elapsed_time(trial_state["timer_trial_elapsed"])
        obj.timer_trial.start()
        return obj

    @rank0
    def save(self):
        save_pickle(self.data, self.fpath_data)


class ArtifactManager:

    dpath_campaign = None
    dpath_setting = None
    dpath_trial = None
    fpath_metadata_trial = None
    dpath_model_final = None
    dpath_eval_final = None
    dpath_model_checkpoint = None
    resuming = False
    dataset = None
    split = None

    # table_eval_group -> (scores set_key, group key, human-readable name); drives map.png and metrics.xlsx
    _TABLE_EVAL_GROUPS = {
        "closed_standard": ("closed_set", "standard", "Closed-Set, Standard"),
        "closed_macro": ("closed_set", "per_class", "Closed-Set, Macro"),
        "full_standard": ("full_set", "standard", "Full-Set, Standard"),
        "full_macro": ("full_set", "per_class", "Full-Set, Macro"),
    }

    @staticmethod
    def set_paths(cfg_train):

        ArtifactManager.dpath_campaign = paths["artifacts"] / cfg_train.campaign
        ArtifactManager.dpath_setting = ArtifactManager.dpath_campaign / "settings" / cfg_train.setting
        ArtifactManager.dataset = cfg_train.dataset
        ArtifactManager.split = cfg_train.split

        trial_name = cfg_train.seed
        ArtifactManager.dpath_trial = ArtifactManager.dpath_setting / cfg_train.dataset / str(trial_name)
        ArtifactManager.fpath_metadata_trial = ArtifactManager.dpath_trial / "trial_metadata.json"

        ArtifactManager.dpath_model_final = ArtifactManager.dpath_trial / "chkpts/final"
        ArtifactManager.dpath_eval_final = ArtifactManager.dpath_trial / "evals/final"
        ArtifactManager.dpath_model_checkpoint = ArtifactManager.dpath_trial / "chkpts/in_progress"

        if ArtifactManager.dpath_trial.exists():
            if (ArtifactManager.dpath_model_checkpoint / "train_state.pt").exists():
                ArtifactManager.resuming = True
            else:
                shutil.rmtree(ArtifactManager.dpath_trial, ignore_errors=True)
                ArtifactManager.resuming = False
        else:
            ArtifactManager.resuming = False

    @staticmethod
    @rank0
    def create_trial_dirs():
        if ArtifactManager.resuming:
            return
        for subdir in ("logs", "chkpts", "chkpts/in_progress", "learning_curves"):
            (ArtifactManager.dpath_trial / subdir).mkdir(parents=True)

    @staticmethod
    @rank0
    def update_campaign_time():

        def format_duration(seconds: float) -> str:
            seconds = int(seconds)
            days, seconds = divmod(seconds, 86400)
            hours, seconds = divmod(seconds, 3600)
            minutes, seconds = divmod(seconds, 60)
            return f"{days}-{hours:02}:{minutes:02}:{seconds:02}"
        
        fpath_pkl = ArtifactManager.dpath_campaign / "time.pkl"
        fpath_json = ArtifactManager.dpath_campaign / "campaign_metadata.json"

        time_data = load_pickle(fpath_pkl)

        time_last_updated = time_data["last_updated"]
        time_curr = time.time()

        time_elapsed = time_data["elapsed"]
        time_elapsed = time_elapsed + (time_curr - time_last_updated)

        time_data["last_updated"] = time_curr
        time_data["elapsed"] = time_elapsed
        save_pickle(time_data, fpath_pkl)
        
        metadata_camp = load_json(ArtifactManager.dpath_campaign / "campaign_metadata.json")
        metadata_camp["duration"] = format_duration(time_elapsed)
        save_json(metadata_camp, fpath_json)

    @staticmethod
    @rank0
    def update_campaign_memory(mem):
        # campaign-level memory = running max across every trial's snapshots (== max across
        # trial-level values, maintained incrementally so it survives an OOM-killed trial)
        fpath_meta = ArtifactManager.dpath_campaign / "campaign_metadata.json"
        metadata_camp = load_json(fpath_meta)
        metadata_camp["memory"] = merge_mem(metadata_camp["memory"], format_mem(mem))
        save_json(metadata_camp, fpath_meta)

    @staticmethod
    @rank0
    def save_metadata_setting(cfg_train):
        
        def clean_metadata(metadata):

            del metadata["campaign"]
            del metadata["setting"]
            del metadata["seed"]
            del metadata["idx_seed"]
            del metadata["dataset"]
            del metadata["split"]

            del metadata["dev"]

            if metadata["loss2"]["mix"] == 0.0:
                del metadata["loss2"]

            # drop inert wting.norm params (per the train.yaml inertness notes): the unit-scale
            # blend (loss / loss.detach()) cancels any per-batch scalar factor on a loss, making
            # norm.agg inert; norm.cls_imb's rescale is such a scalar under multiplicative aggs,
            # cancelled by norm.agg or unit-scaling. loss2 is already gone when mix = 0.0, under
            # which mix_unit_scale never applies
            unit_scaled = "loss2" in metadata and metadata["loss2"]["mix_unit_scale"]
            for key in ("loss", "loss2"):
                if key in metadata:
                    wting = metadata[key]["wting"]
                    if wting["agg"] in ("prod", "geo_mean") and (wting["norm"]["agg"] or unit_scaled):
                        del wting["norm"]["cls_imb"]
                    if unit_scaled:
                        del wting["norm"]["agg"]

        fpath_meta = ArtifactManager.dpath_setting / "setting_metadata.json"
        metadata = asdict(cfg_train)
        clean_metadata(metadata)
        if fpath_meta.exists():
            metadata_loaded = load_json(fpath_meta)
            assert metadata == metadata_loaded, "Setting params changed!"
        else:
            save_json(metadata, fpath_meta)

    @staticmethod
    def _get_trial_runtime_data(data: TrialData, idx_epoch: int, time_tracker: TimeTracker):

        def fmt(seconds):
            return f"{seconds:.2f}" if seconds is not None else None

        def mean_bucket(name):
            return {"mean": fmt(time_tracker.mean(name)), "n": time_tracker.n(name)}

        # train mean keyed on idx_epoch (epochs *started*) to match its per-epoch cadence; shows
        # "0.00" between the first epoch's start and finish, as before.
        if idx_epoch > 0:
            mean_time_train = fmt(time_tracker.mean("train") or 0.0)
        else:
            mean_time_train = None

        time_trial = data.timer_trial.get_elapsed_time()
        # remainder of the trial wall-clock not attributed to a tracked bucket -- checkpoint I/O,
        # sync/barriers, and any in-progress epoch's train time not yet folded into the train mean
        time_other = time_trial - time_tracker.attributed()

        runtime_data = {
            "train": {"mean": mean_time_train, "n": idx_epoch},
            "eval": mean_bucket("eval"),
            "viz_compute": mean_bucket("viz_compute"),
            "other": fmt(time_other),
            "trial": fmt(time_trial),
        }

        return runtime_data

    @staticmethod
    @rank0
    def save_metadata_trial(data: TrialData, idx_epoch: int, time_tracker: TimeTracker, n_samps_seen: int, sample_volume: int, mem, init_flag=False):
        runtime_data = ArtifactManager._get_trial_runtime_data(data, idx_epoch, time_tracker)
        progress_data = {"n_samps_seen": n_samps_seen, "sample_volume": sample_volume}
        now = datetime.now(timezone.utc).strftime("%Y-%m-%dT%H:%M:%SZ")
        if init_flag:
            metadata_trial = {
                "dataset": ArtifactManager.dataset,
                "split": ArtifactManager.split,
                "runtime": runtime_data,
                "progress": progress_data,
                "memory": format_mem(mem),
                "datetime_start": now,
                "datetime_last_seen": now,
                "complete": False,
            }
        else:
            metadata_trial = load_json(ArtifactManager.fpath_metadata_trial)
            metadata_trial["runtime"] = runtime_data
            metadata_trial["progress"] = progress_data
            metadata_trial["memory"] = merge_mem(metadata_trial["memory"], format_mem(mem))
            metadata_trial["datetime_last_seen"] = now
        save_json(metadata_trial, ArtifactManager.fpath_metadata_trial)

    @staticmethod
    @rank0
    def save_eval_data(dpath_model, eval_metrics, n_samps_seen_chkpt, n_samps_seen):
        dpath_model.mkdir(parents=True, exist_ok=True)
        fpath_meta = dpath_model / "metrics.json"
        metadata = {
            **format_scores(eval_metrics),
            "n_samps_seen": f"{n_samps_seen_chkpt:,}/{n_samps_seen:,}",
        }
        save_json(metadata, fpath_meta)

    @staticmethod
    def _spread(nums, spread_type):
        spread = nums.std(ddof=1)
        return spread / np.sqrt(len(nums)) if spread_type == "ste" else spread

    @staticmethod
    def _aggregate_metric_stats(values, spread_type, percent=True):
        first = values[0]
        if first is None:  # leaf is None for every trial (e.g. loss_raw["ood"] is never computed)
            return None
        if isinstance(first, dict):
            return {
                k: ArtifactManager._aggregate_metric_stats(
                    [v[k] for v in values], spread_type, percent=percent and k not in ("loss_raw", "sim", "targ")
                )
                for k in first
            }
        if len(values) == 1:
            return values[0]
        nums = np.array([float(v) for v in values])
        mean = nums.mean()
        spread = ArtifactManager._spread(nums, spread_type)
        if percent:
            return f"{mean * 100:.2f} ± {spread * 100:.2f}"
        return f"{mean:.4f} ± {spread:.4f}"

    @staticmethod
    def _listview_metric_stats(values, percent=True):
        first = values[0]
        if first is None:  # leaf is None for every trial (mirrors _aggregate_metric_stats)
            return None
        if isinstance(first, dict):
            return {
                k: ArtifactManager._listview_metric_stats(
                    [v[k] for v in values], percent=percent and k not in ("loss_raw", "sim", "targ")
                )
                for k in first
            }
        if percent:
            return [f"{float(v) * 100:.2f}" for v in values]
        return [f"{float(v):.4f}" for v in values]

    @staticmethod
    @rank0
    def update_metric_stats(spread_type):
        dpath_dataset = ArtifactManager.dpath_setting / ArtifactManager.dataset
        metric_dicts = []
        for dpath_trial in sorted(dpath_dataset.iterdir()):
            # a written final-eval metrics file is the signal a trial finished; the `complete` flag is
            # marked later (campaign_runner, after a clean exit) so it can't gate this aggregation
            fpath_metrics = dpath_trial / "evals/final/metrics.json"
            if not fpath_metrics.exists():
                continue
            metrics = load_json(fpath_metrics)
            metrics.pop("n_samps_seen", None)
            metric_dicts.append(metrics)

        if not metric_dicts:
            return

        n_trials = len(metric_dicts)
        stats = {
            "n_trials": n_trials,
            **ArtifactManager._aggregate_metric_stats(metric_dicts, spread_type),
        }
        listview = {
            "n_trials": n_trials,
            **ArtifactManager._listview_metric_stats(metric_dicts),
        }
        dpath_stats = dpath_dataset / "stats"
        dpath_stats.mkdir(parents=True, exist_ok=True)
        save_json(stats, dpath_stats / "metrics.json")
        save_json_listview(listview, dpath_stats / "metrics_listview.json")

    @staticmethod
    def _stats_table_grid(labels, setting_score_maps, spread_type):
        """Build a composite-score table's cell grid from [(setting, [score dict per completed
        trial]), ...]: a header row of 'Setting' + one column per label in `labels`, then one row
        per setting -- '<setting> (n_trials)' + each label's cell (read from the score dicts by
        its lowercased key): '-' (0 trials), 'XX.XX' (1 trial, mean), or 'XX.XX ± XX.XX'
        (>1 trial, mean ± spread)."""
        grid = [["Setting", *labels]]
        for setting, score_maps in setting_score_maps:
            row = [f"{setting} ({len(score_maps)})"]
            for label in labels:
                nums = np.array([float(score_map[label.lower()]) for score_map in score_maps]) * 100
                if len(nums) == 0:
                    row.append("-")
                elif len(nums) == 1:
                    row.append(f"{nums[0]:.2f}")
                else:
                    row.append(f"{nums.mean():.2f} ± {ArtifactManager._spread(nums, spread_type):.2f}")
            grid.append(row)
        return grid

    @staticmethod
    def _collect_comps(settings, datasets, set_key, grp):
        """Each (setting, dataset)'s completed-trial score maps, keyed by trial seed (the trial dir
        name; empty dict -> no trials yet): per trial a {'map': ..., 'acc': ...} pair of flat
        label->score dicts merging the comp scores with the per-partition primitives ('id i2t' ...
        'ood t2i'), so table labels map to keys by lowercasing. A written final-eval metrics file is
        the completion signal, same as update_metric_stats."""
        comps_by = {}
        for setting in settings:
            for dataset in datasets:
                dpath_dataset = ArtifactManager.dpath_campaign / "settings" / setting / dataset
                comps = {}
                if dpath_dataset.exists():
                    for dpath_trial in sorted(dpath_dataset.iterdir()):
                        fpath_metrics = dpath_trial / "evals/final/metrics.json"
                        if fpath_metrics.exists():
                            scores_grp = load_json(fpath_metrics)["scores"][set_key][grp]
                            comps[dpath_trial.name] = {
                                "map": {**scores_grp["comp"]["map"],
                                        **{f"{p} {m}": scores_grp[p]["map"][m] for p in ("id", "ood") for m in ("i2t", "i2i", "t2i")}},
                                "acc": {**scores_grp["comp"]["acc"],
                                        **{f"{p} i2t": scores_grp[p]["acc"]["i2t"] for p in ("id", "ood")}},
                            }
                comps_by[(setting, dataset)] = comps
        return comps_by

    @staticmethod
    def _score_labels(prim_scores):
        """(mAP labels, acc labels) for the stats tables; prim_scores appends the per-partition
        primitive score columns (ID/OOD x modality) to each."""
        map_labels = ("All", "ID", "OOD", "I2T", "I2I", "T2I")
        acc_labels = ("I2T",)
        if prim_scores:
            map_labels += ("ID I2T", "ID I2I", "ID T2I", "OOD I2T", "OOD I2I", "OOD T2I")
            acc_labels += ("ID I2T", "OOD I2T")
        return map_labels, acc_labels

    @staticmethod
    def _cross_dataset_means(settings, datasets, comps_by, score_key, labels):
        """xmeans[(setting, label)]: arithmetic mean, across datasets with completed trials, of that
        setting/label's per-dataset mean comp score (percent), read from comp[score_key][label.lower()]
        (score_key: 'map' or 'acc'); None when no dataset has trials."""

        def dataset_means(setting, label):
            key = label.lower()
            return [np.mean([float(comp[score_key][key]) for comp in comps_by[(setting, dataset)].values()]) * 100
                    for dataset in datasets if comps_by[(setting, dataset)]]

        xmeans = {}
        for setting in settings:
            for label in labels:
                vals = dataset_means(setting, label)
                xmeans[(setting, label)] = np.mean(vals) if vals else None
        return xmeans

    @staticmethod
    def _order_settings(settings, xmeans, label):
        # order setting rows by the mean for `label`, descending (ties keep campaign order); callers
        # filter out settings with no completed trials before ordering, so every mean is numeric
        return sorted(settings, key=lambda s: xmeans[(s, label)], reverse=True)

    @staticmethod
    def _col_styles(grid, bold_high):
        """Per-column data-cell styling for one rendered table, shared by the png and xlsx tables:
        styles[c] for each score-label column c -- row -> mean for numeric cells ('-' skipped), the
        bold-winner rows (highest mean, ties included; empty unless bold_high), and the column's
        min/max mean for scaled heatmaps."""
        styles = {}
        for c in range(1, len(grid[0])):
            means = {r: float(grid[r][c].split(" ± ")[0]) for r in range(1, len(grid)) if grid[r][c] != "-"}
            winners = set()
            if bold_high and means:
                top = max(means.values())
                winners = {r for r, m in means.items() if m == top}
            col_min = min(means.values()) if means else 0.0
            col_max = max(means.values()) if means else 0.0
            styles[c] = (means, winners, col_min, col_max)
        return styles

    @staticmethod
    def _heat_hex(heatmap, mean, col_min, col_max):
        """Heatmap cell color as 'RRGGBB': linear white (#ffffff) -> #ff5533 interpolation -- 'fixed'
        maps a fixed 0.00 -> 100.00, 'scaled' maps the column's min -> max (a single-value or
        all-equal column -> lowest color)."""
        if heatmap == "fixed":
            frac = mean / 100.0
        elif col_max > col_min:  # scaled across the column's data cells
            frac = (mean - col_min) / (col_max - col_min)
        else:  # scaled but column has one value (or all equal) -> lowest color
            frac = 0.0
        frac = max(0.0, min(1.0, frac))
        g = round(255 - (255 - 0x55) * frac)
        b = round(255 - (255 - 0x33) * frac)
        return f"FF{g:02X}{b:02X}"

    @staticmethod
    def _render_stats_table(grid, title, fpath, bold_high, heatmap):
        fig, ax = plt.subplots(figsize=(1.2 + 1.5 * (len(grid[0]) - 1), 0.7 + 0.3 * len(grid)))
        ax.axis("off")
        ax.set_title(title, fontsize=11, pad=12)
        table = ax.table(cellText=grid, loc="center", cellLoc="center")
        table.auto_set_font_size(False)
        table.set_fontsize(9)
        table.scale(1, 1.5)
        table.auto_set_column_width(list(range(len(grid[0]))))
        styles = ArtifactManager._col_styles(grid, bold_high)
        for (row, col), cell in table.get_celld().items():
            if row == 0 or col == 0:
                cell.set_text_props(fontweight="bold")
                cell.set_facecolor("#eaeaea")
                continue
            means, winners, col_min, col_max = styles[col]
            if row in winners:
                cell.set_text_props(fontweight="bold")
            if heatmap and row in means:
                cell.set_facecolor(f"#{ArtifactManager._heat_hex(heatmap, means[row], col_min, col_max)}")
        fig.savefig(fpath, dpi=200, bbox_inches="tight")
        plt.close(fig)

    @staticmethod
    @rank0
    def update_stats_tables(table_eval_group, spread_type, bold_high, ordered, heatmap, prim_scores):
        """Render the campaign-level composite-score summary tables for this trial's dataset to
        artifacts/<campaign>/stats/<dataset>/map.png (comp mAP: All/ID/OOD/I2T/I2I/T2I score
        columns) and acc.png (comp I2T accuracy: single I2T column) -- prim_scores appends the
        per-partition primitive score columns (ID/OOD x modality) to both: one row per setting with >= 1
        completed trial in this dataset (settings without local trials are omitted -- no blank rows
        in the pngs), stats aggregated across each setting's completed trials. bold_high/ordered/
        heatmap style the
        tables the same way as metrics.xlsx: bold_high bolds each score column's highest-mean cell
        (ties included; '-' cells ignored), ordered orders each table's setting rows by its own
        metric's mean over THIS dataset's completed trials (map.png by the mAP 'All' column,
        acc.png by the acc 'I2T' column) -- localized per dataset, independent of the cross-dataset
        order used in metrics.xlsx -- heatmap shades score cells white->#ff5533 (None/scaled/fixed
        as in update_metrics_xlsx). Re-rendered at each trial completion."""
        set_key, grp, group_name = ArtifactManager._TABLE_EVAL_GROUPS[table_eval_group]

        settings = load_json(ArtifactManager.dpath_campaign / "campaign_metadata.json")["settings"]
        dataset = ArtifactManager.dataset
        comps_by = ArtifactManager._collect_comps(settings, (dataset,), set_key, grp)
        # a setting gets a row only once it has >= 1 completed trial in THIS dataset (no blank rows)
        settings = [s for s in settings if comps_by[(s, dataset)]]

        def ordered_settings(score_key, label):
            # localized order: this dataset's per-setting trial means (single-dataset degenerate
            # case of _cross_dataset_means), not the cross-dataset means backing the xlsx order
            means = ArtifactManager._cross_dataset_means(settings, (dataset,), comps_by, score_key, (label,))
            return ArtifactManager._order_settings(settings, means, label)

        settings_map = ordered_settings("map", "All") if ordered else settings
        settings_acc = ordered_settings("acc", "I2T") if ordered else settings
        map_labels, acc_labels = ArtifactManager._score_labels(prim_scores)

        title_suffix = f" -- {DATASET_ALIAS2NAME[dataset]} ({group_name})"
        dpath_stats = ArtifactManager.dpath_campaign / "stats" / dataset
        dpath_stats.mkdir(parents=True, exist_ok=True)

        grid_map = ArtifactManager._stats_table_grid(
            map_labels,
            [(setting, [comp["map"] for comp in comps_by[(setting, dataset)].values()]) for setting in settings_map],
            spread_type,
        )
        ArtifactManager._render_stats_table(grid_map, f"Composite mAP{title_suffix}", dpath_stats / "map.png", bold_high, heatmap)

        grid_acc = ArtifactManager._stats_table_grid(
            acc_labels,
            [(setting, [comp["acc"] for comp in comps_by[(setting, dataset)].values()]) for setting in settings_acc],
            spread_type,
        )
        ArtifactManager._render_stats_table(grid_acc, f"Composite I2T Accuracy{title_suffix}", dpath_stats / "acc.png", bold_high, heatmap)

    @staticmethod
    @rank0
    def update_metrics_xlsx(table_eval_group, spread_type, bold_high, ordered, heatmap, prim_scores, baseline_overrides):
        """Write artifacts/<campaign>/stats/metrics.xlsx: two sheets, 'Composite mAP' (comp map scores,
        All/ID/OOD/I2T/I2I/T2I score columns) and 'Composite I2T Accuracy' (comp acc, single I2T
        column) -- prim_scores appends the per-partition primitive score columns (ID/OOD x modality)
        to both sheets' tables, and splits each mAP-sheet table banner into the title cell (first
        column, unmerged) plus grey merged 'Composite Scores' / 'Primitive Scores' group headers over
        their column groups (the accuracy sheet keeps full-width merged title banners). Each sheet
        opens with a bold '<repo-parent-dir> - <campaign>' title cell (e.g.
        'bc_dev - dev') and a blank row, then stacks one table per campaign dataset vertically -- a bold left-aligned
        title banner, then a table of header row 'Setting' + one column per score label and one
        '<setting> (n_trials)' row per setting, then a blank spacer row before the next dataset -- with
        the always-shown 'Mean' summary table at the bottom: one row per setting, each cell the
        arithmetic mean, across datasets with completed trials, of that setting/label's per-dataset
        mean (a point value, no spread). Cells are '-' (0 trials), 'XX.XX' (1 trial, mean) or 'XX.XX ± XX.XX'
        (>1 trial, mean ± spread), aggregated across each setting's completed trials'
        scores.<set_key>.<grp>.comp for the eval group selected by table_eval_group. A setting gets
        rows only once it has >= 1 completed trial in some dataset -- it then appears in every table of
        both sheets, with blank '-' rows in dataset tables lacking its trials; settings with no
        completed trials anywhere are omitted entirely. When bold_high is
        True, the highest-mean setting cell in each score column is bolded (ties included; '-' cells
        ignored). When ordered is True, each sheet's setting rows are ordered by its own metric's
        Mean-table first score column -- 'All' for mAP, 'I2T' for accuracy -- descending, settings with
        no completed trials anywhere last; when False, rows keep the fixed campaign_metadata order.
        Within a sheet one row order is shared across all tables, but the two sheets' orders may differ.
        heatmap shades each score cell white->#ff5533 by intensity: None leaves cells unshaded; 'scaled'
        maps each score column's min->max to white->#ff5533; 'fixed' maps a fixed 0.00->100.00 to
        white->#ff5533. '-' cells are never shaded. To the right of this aggregate block sit per-seed
        blocks (one blank separator column apart): a 'seed <seed>' label in the campaign-banner row,
        then the per-dataset tables only (no Mean summary) built from that seed's trials alone,
        sitting in the same rows as the aggregate block's dataset tables -- plain setting labels (no
        trial counts), single-trial 'XX.XX' cells, '-' where that seed's trial hasn't completed --
        sharing the aggregate block's setting rows/order. baseline_overrides adds a 'Baseline
        Overrides' config table to each sheet, in a left column band that the score blocks shift
        right past (one blank separator column between), vertically aligned with the aggregate
        block's bottom Mean table so the Mean's Setting column labels its rows: one column per param
        overridden in the campaign's baseline_overrides (union of the settings' overrides.json keys,
        first-seen order), each cell the setting's effective value resolved from its
        setting_metadata.json -- '-' when the param is absent there, the signal that it is inert
        under that configuration (e.g. loss2.* with loss2.mix 0.0). This table gets no
        winner-bold/heatmap styling. Column widths hug each column's longest header/data cell
        (banner/label text overflows); blank separator columns get a small ~square width. Regenerated
        at each trial completion."""
        set_key, grp, _ = ArtifactManager._TABLE_EVAL_GROUPS[table_eval_group]
        metadata = load_json(ArtifactManager.dpath_campaign / "campaign_metadata.json")
        settings, datasets = metadata["settings"], metadata["datasets"]

        comps_by = ArtifactManager._collect_comps(settings, datasets, set_key, grp)
        # a setting gets rows only once it has >= 1 completed trial in some dataset; it then appears in
        # every dataset table (blank '-' row where that dataset has no trials for it yet)
        settings = [s for s in settings if any(comps_by[(s, dataset)] for dataset in datasets)]
        seeds = sorted({seed for comps in comps_by.values() for seed in comps}, key=int)

        if baseline_overrides:
            dpath_settings = ArtifactManager.dpath_campaign / "settings"
            okeys = []  # union of overridden params, first-seen order across settings (campaign order)
            for s in settings:
                for key in load_json(dpath_settings / s / "overrides.json"):
                    if key not in okeys:
                        okeys.append(key)
            metadata_by = {s: load_json(dpath_settings / s / "setting_metadata.json") for s in settings}

            def override_value(setting, key):
                # a param absent from the setting's metadata is inert under that configuration -> '-'
                node = metadata_by[setting]
                for part in key.split("."):
                    if not isinstance(node, dict) or part not in node:
                        return "-"
                    node = node[part]
                return str(node)

        def build_blocks(score_key, labels):
            """The sheet's blocks, left to right: (label, [(title, cell grid), ...]) -- the aggregate
            block (label None): one table per campaign dataset, then the always-shown 'Mean'
            cross-dataset summary table at the bottom; then one block per completed seed (label
            'seed <seed>'): the per-dataset tables only (no Mean summary), built from that seed's
            trials alone -- plain setting labels (no trial counts), single-trial 'XX.XX' cells, '-'
            where that seed's trial hasn't completed. Setting rows are shared across all blocks --
            when ordered, pinned to the aggregate Mean-table's first score column (labels[0]),
            descending. Returns (blocks, ogrid): ogrid is the 'Baseline Overrides' grid (param-name
            header + one value row per setting in this sheet's row order), or None when disabled or
            nothing is overridden."""
            xmeans = ArtifactManager._cross_dataset_means(settings, datasets, comps_by, score_key, labels)
            rows = ArtifactManager._order_settings(settings, xmeans, labels[0]) if ordered else settings

            tables = []
            for dataset in datasets:
                grid = ArtifactManager._stats_table_grid(
                    labels,
                    [(setting, [comp[score_key] for comp in comps_by[(setting, dataset)].values()]) for setting in rows],
                    spread_type,
                )
                tables.append((DATASET_ALIAS2NAME[dataset], grid))
            xgrid = [["Setting", *labels]]
            for s in rows:
                xgrid.append([s] + [f"{xmeans[(s, label)]:.2f}" for label in labels])
            tables.append(("Mean", xgrid))
            blocks = [(None, tables)]

            for seed in seeds:
                stables = []
                for dataset in datasets:
                    grid = [["Setting", *labels]]
                    for s in rows:
                        comp = comps_by[(s, dataset)].get(seed)
                        grid.append([s] + ["-" if comp is None else f"{float(comp[score_key][label.lower()]) * 100:.2f}"
                                           for label in labels])
                    stables.append((DATASET_ALIAS2NAME[dataset], grid))
                blocks.append((f"seed {seed}", stables))

            ogrid = None
            if baseline_overrides and okeys:
                # header of param names only -- no setting column; the rows align with (and are
                # labeled by) the aggregate Mean table's setting rows
                ogrid = [list(okeys)]
                for s in rows:
                    ogrid.append([override_value(s, key) for key in okeys])
            return blocks, ogrid

        bold = Font(bold=True)
        center = Alignment(horizontal="center", vertical="center")
        left = Alignment(horizontal="left", vertical="center")
        header_fill = PatternFill("solid", fgColor="EAEAEA")
        thin = Side(style="thin", color="000000")
        border = Border(left=thin, right=thin, top=thin, bottom=thin)

        def write_sheet(ws, blocks, groups, ogrid):
            """groups: None -> each table's banner is its title merged across the full table width;
            else [(group_title, n_group_cols), ...] -> the title sits unmerged in the block's first
            column, followed by one grey merged group-header cell per group (e.g. 'Composite Scores'
            over the composite columns, 'Primitive Scores' over the primitive columns). ogrid: None,
            or the 'Baseline Overrides' grid (param-name header + one value row per setting, sheet
            row order) -- rendered in a left column band (one blank separator column before the
            score blocks, which all shift right past it), vertically aligned with the aggregate
            block's bottom Mean table so the Mean's Setting column labels its rows; config cells get
            no winner-bold/heatmap styling."""
            n_cols = len(blocks[0][1][0][1][0])  # corner + one col per score label (same for every grid in the sheet)
            widths = {}  # col idx -> longest header/data cell text (banner/label cells overflow instead)

            campaign = ws.cell(row=1, column=1, value=f"{paths['root'].parent.name} - {ArtifactManager.dpath_campaign.name}")
            campaign.font = bold
            campaign.alignment = left

            offset = len(ogrid[0]) + 1 if ogrid else 0  # overrides band + its separator column
            for i_block, (block_label, tables) in enumerate(blocks):
                col0 = 1 + offset + i_block * (n_cols + 1)  # blocks side by side, one blank separator column apart
                if block_label is not None:
                    label_cell = ws.cell(row=1, column=col0, value=block_label)  # campaign-banner row, atop the block
                    label_cell.font = bold
                    label_cell.alignment = left
                # banner row + blank row above the tables; dataset tables lead in every block, so they
                # sit in the same rows across blocks (only the aggregate has the trailing Mean table)
                row = 3
                for title_text, grid in tables:
                    title = ws.cell(row=row, column=col0, value=title_text)
                    title.font = bold
                    title.alignment = left
                    for c in range(col0, col0 + n_cols):  # border every cell of the banner so the merged ranges' edges all render
                        ws.cell(row=row, column=c).border = border
                    if groups is None:
                        ws.merge_cells(start_row=row, start_column=col0, end_row=row, end_column=col0 + n_cols - 1)
                    else:
                        widths[col0] = max(widths.get(col0, 0), len(title_text))  # unmerged title must fit its column
                        gcol = col0 + 1
                        for group_title, n_group in groups:
                            for c in range(gcol, gcol + n_group):  # fill every cell so the merged range renders grey
                                ws.cell(row=row, column=c).fill = header_fill
                            gcell = ws.cell(row=row, column=gcol, value=group_title)
                            gcell.font = bold
                            gcell.alignment = center
                            ws.merge_cells(start_row=row, start_column=gcol, end_row=row, end_column=gcol + n_group - 1)
                            gcol += n_group
                    row += 1

                    styles = ArtifactManager._col_styles(grid, bold_high)
                    for r, grid_row in enumerate(grid):
                        for c, val in enumerate(grid_row):
                            cell = ws.cell(row=row, column=col0 + c, value=val)
                            cell.alignment = center
                            cell.border = border
                            widths[col0 + c] = max(widths.get(col0 + c, 0), len(val))
                            if r == 0 or c == 0:
                                cell.font = bold
                                cell.fill = header_fill
                                continue
                            means, winners, col_min, col_max = styles[c]
                            if r in winners:
                                cell.font = bold
                            if heatmap and r in means:
                                cell.fill = PatternFill("solid", fgColor=ArtifactManager._heat_hex(heatmap, means[r], col_min, col_max))
                        row += 1
                    row += 1  # blank spacer row between tables

            if ogrid:
                o_cols = len(ogrid[0])
                # banner row of the aggregate block's bottom Mean table (dataset tables precede it)
                row = 3 + sum(len(grid) + 2 for _, grid in blocks[0][1][:-1])
                title = ws.cell(row=row, column=1, value="Baseline Overrides")
                title.font = bold
                title.alignment = left
                for c in range(1, o_cols + 1):
                    ws.cell(row=row, column=c).border = border
                ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=o_cols)
                row += 1
                for r, grid_row in enumerate(ogrid):
                    for c, val in enumerate(grid_row):
                        cell = ws.cell(row=row, column=1 + c, value=val)
                        cell.alignment = center
                        cell.border = border
                        widths[1 + c] = max(widths.get(1 + c, 0), len(val))
                        if r == 0:  # param-name header row
                            cell.font = bold
                            cell.fill = header_fill
                    row += 1

            for c in range(1, max(widths) + 1):
                # snug fit to each column's longest cell; blank separator columns get a small ~square width
                ws.column_dimensions[get_column_letter(c)].width = widths[c] + 2 if c in widths else 3

        map_labels, acc_labels = ArtifactManager._score_labels(prim_scores)
        # with prim_scores, mAP-sheet banners split into title + 'Composite Scores'/'Primitive Scores'
        # group headers; the accuracy sheet keeps full-width title banners
        map_groups = [("Composite Scores", 6), ("Primitive Scores", 6)] if prim_scores else None
        wb = Workbook()
        ws_map = wb.active
        ws_map.title = "Composite mAP"
        map_blocks, map_ogrid = build_blocks("map", map_labels)
        write_sheet(ws_map, map_blocks, map_groups, map_ogrid)
        acc_blocks, acc_ogrid = build_blocks("acc", acc_labels)
        write_sheet(wb.create_sheet("Composite I2T Accuracy"), acc_blocks, None, acc_ogrid)

        dpath_stats = ArtifactManager.dpath_campaign / "stats"
        dpath_stats.mkdir(parents=True, exist_ok=True)
        wb.save(dpath_stats / "metrics.xlsx")

    @staticmethod
    def base_eval_cache_fpath(cfg_train):
        # one pickle per combo in a flat dir, named by the serialized combo key -- a save touches
        # only its own combo's file (no shared-file rewrite) and a load reads only its own
        fname = "__".join(str(c) for c in ArtifactManager.base_eval_key(cfg_train)) + ".pkl"
        return paths["root"] / "base_eval_cache" / fname

    @staticmethod
    def base_eval_key(cfg_train):
        """Combo key for one base-eval reading: the config settings that determine the base model's
        eval output. Numerics-level knobs (hw mixed_prec, t-SNE perplexity) are
        deliberately not keyed. Family-inert components are normalized to None so equivalent configs
        share one entry: non_causal is CLIP-only, vis_proj is SigLIP-only, and seed only enters
        through the random init of a linear/mlp vis_proj head."""
        from models import CLIP_MODELS, SIGLIP_MODELS  # local: models pulls open_clip/transformers, too heavy for module import
        model_type = cfg_train.arch["model_type"]
        non_causal = cfg_train.arch["clip"]["non_causal"] if model_type in CLIP_MODELS else None
        vis_proj = cfg_train.arch["siglip"]["vis_proj"] if model_type in SIGLIP_MODELS else None
        seed = cfg_train.seed if vis_proj is not None else None
        return (
            model_type,
            cfg_train.img_norm,
            cfg_train.dataset,
            cfg_train.split,
            non_causal,
            cfg_train.text_template["eval"],
            vis_proj,
            seed,
        )

    @staticmethod
    @rank0
    def load_base_eval_cache(cfg_train, require_projections, require_embs):
        # @rank0: a concurrent same-combo campaign can create/replace this combo's file at any
        # moment, so independent per-rank reads could disagree on hit/miss; rank 0 alone reads and
        # the caller broadcasts the decision. Entries carry only what the caching trial computed
        # (metrics always; projections for viz trials; embs for pooled trials): a trial must read
        # an entry missing a piece it needs as a miss (recompute, overwriting the entry with the
        # richer version) rather than trip _write_base_eval on the missing piece downstream.
        # Leaner entries stay valid hits for trials that don't need the missing pieces.
        fpath = ArtifactManager.base_eval_cache_fpath(cfg_train)
        if not fpath.exists():
            return None
        entry = load_pickle(fpath)
        if require_projections and entry["projections"] is None:
            return None
        if require_embs and entry["embs"] is None:
            return None
        return entry

    @staticmethod
    @rank0
    def save_base_eval_cache(cfg_train, eval_metrics):
        """Write this combo's entry to its own cache file and return the entry. The npz arrays are
        ingested from this trial's evals/_base/, where compute_projections just wrote them
        (projections absent for non-viz trials, embs for non-pooled trials). Written via temp file +
        atomic replace: concurrent same-combo campaigns overwrite each other with equivalent entries,
        and readers never see a torn file; other combos' files are untouched."""
        entry = {
            "metrics": {k: format_scores(v) for k, v in eval_metrics.items() if k not in ("loss_raw", "sim", "targ")},
            "projections": None,
            "embs": None,
        }
        dpath_base = ArtifactManager.dpath_trial / "evals" / "_base"
        for name in ("projections", "embs"):
            fpath_npz = dpath_base / f"{name}.npz"
            if fpath_npz.exists():
                entry[name] = dict(np.load(fpath_npz))
        fpath = ArtifactManager.base_eval_cache_fpath(cfg_train)
        fpath.parent.mkdir(parents=True, exist_ok=True)
        fpath_tmp = fpath.with_name(f"{fpath.name}.{os.uname().nodename}.{os.getpid()}.tmp")  # node+pid: campaigns on different nodes can share a pid
        save_pickle(entry, fpath_tmp)
        fpath_tmp.replace(fpath)
        return entry

    @staticmethod
    @rank0
    def save_train_state(train_pipe, idx_batch):
        state = {
            "model": train_pipe.modelw._unwrapped_model.state_dict(),
            "norm_mean": train_pipe.modelw.norm_mean,
            "norm_std": train_pipe.modelw.norm_std,
            "optimizer": train_pipe.opt.state_dict(),
            "lr_sched": train_pipe.lr_sched.state_dict(),
            "n_samps_seen": train_pipe.n_samps_seen,
            "n_batches_seen": train_pipe.n_batches_seen,
            "idx_epoch": train_pipe.idx_epoch,
            "idx_batch": idx_batch,
            "chkpt_thresh": train_pipe.chkpt_thresh,
            "times": train_pipe.time_tracker.state_dict(),
        }
        torch.save(state, ArtifactManager.dpath_model_checkpoint / "train_state.pt")

    @staticmethod
    def save_rng_states(rank):
        rng_state = {
            "rng_cpu": torch.get_rng_state(),
            "rng_cuda": torch.cuda.get_rng_state_all(),
            "rng_numpy": np.random.get_state(),
            "rng_random": random.getstate(),
        }
        torch.save(rng_state, ArtifactManager.dpath_model_checkpoint / f"rng_state_rank{rank}.pt")

    @staticmethod
    @rank0
    def save_trial_state(data):
        state = {
            "n_evals": data.n_evals,
            "timer_trial_elapsed": data.timer_trial.get_elapsed_time(),
        }
        save_pickle(state, ArtifactManager.dpath_model_checkpoint / "trial_state.pkl")

    @staticmethod
    def load_train_state():
        return torch.load(
            ArtifactManager.dpath_model_checkpoint / "train_state.pt",
            map_location="cpu",
            weights_only=False,
        )

    @staticmethod
    def load_rng_state(rank):
        return torch.load(
            ArtifactManager.dpath_model_checkpoint / f"rng_state_rank{rank}.pt",
            map_location="cpu",
            weights_only=False,
        )

    @staticmethod
    def load_trial_state():
        return load_pickle(ArtifactManager.dpath_model_checkpoint / "trial_state.pkl")


def _samples_seen_tick_formatter(value, _pos):
    return f"{value / 1_000_000:g}"

@rank0
def plot_metrics(
        data_tracker, 
        dpath_trial,
        fontsize_axes=12, 
        fontsize_ticks=8, 
        fontsize_legend=8,
        subplot_border_width=1,
        figsize=(10, 16),
        height_ratios=[2, 2, 2, 2, 2, 1, 1, 1],
    ):
    data = data_tracker.data
    data_epoch = data["epoch"]
    data_eval = data["eval"]
    title_suffix = f" -- {ArtifactManager.dpath_setting.name}, {DATASET_ALIAS2NAME[ArtifactManager.dataset]}"

    # eval panels (retrieval / n-shot / accuracy) are populated only when eval ran;
    # train panels (loss / grad norm / lr) plot whenever train data is present (e.g. train_pt=trainval).
    partitions = [k for k in data_eval.get("scores", {}).get("closed_set", {}).get("standard", {}).keys() if k != "comp"]

    x_eval = data_eval["n_samps_seen"]
    x_train = data_epoch["n_samps_seen"]

    bucket_partition = "id" if "id" in partitions else None
    bucket_comp_keys_standard = [
        key for key in data_eval.get("scores", {}).get("closed_set", {}).get("standard", {}).get(bucket_partition, {}).get("map", {}).get("n-shot", {}).keys()
    ]
    bucket_comp_keys_full_set = [
        key
        for key in data_eval.get("scores", {}).get("full_set", {}).get("standard", {}).get(bucket_partition, {}).get("map", {}).get("n-shot", {}).keys()
    ]

    plot_composite_metrics(
        data_epoch,
        data_eval,
        x_train,
        x_eval,
        dpath_trial,
        partitions,
        bucket_partition,
        bucket_comp_keys_standard,
        fontsize_axes,
        fontsize_ticks,
        fontsize_legend,
        subplot_border_width,
        figsize,
        height_ratios,
        partition_metric_group="standard",
        full_set=False,
        retrieval_ylabel="mAP Scores",
        accuracy_ylabel="I2T Accuracy",
        nshot_accuracy_ylabel="n-shot Accuracy (ID)",
        plot_title=f"Train Metrics{title_suffix}",
        output_filename="closed_standard.png",
    )

    plot_composite_metrics(
        data_epoch,
        data_eval,
        x_train,
        x_eval,
        dpath_trial,
        partitions,
        bucket_partition,
        bucket_comp_keys_standard,
        fontsize_axes,
        fontsize_ticks,
        fontsize_legend,
        subplot_border_width,
        figsize,
        height_ratios,
        partition_metric_group="per_class",
        full_set=False,
        retrieval_ylabel="Macro mAP Scores",
        accuracy_ylabel="I2T Per-Class Accuracy",
        nshot_accuracy_ylabel="n-shot Per-Class\nAccuracy (ID)",
        plot_title=f"Train Metrics (Macro){title_suffix}",
        output_filename="closed_macro.png",
    )

    plot_composite_metrics(
        data_epoch,
        data_eval,
        x_train,
        x_eval,
        dpath_trial,
        partitions,
        bucket_partition,
        bucket_comp_keys_full_set,
        fontsize_axes,
        fontsize_ticks,
        fontsize_legend,
        subplot_border_width,
        figsize,
        height_ratios,
        partition_metric_group="standard",
        full_set=True,
        retrieval_ylabel="Full-Set mAP Scores",
        accuracy_ylabel="Full-Set I2T Accuracy",
        nshot_accuracy_ylabel="Full-Set n-shot Accuracy (ID)",
        plot_title=f"Train Metrics (Full-Set){title_suffix}",
        output_filename="full_standard.png",
    )

    plot_composite_metrics(
        data_epoch,
        data_eval,
        x_train,
        x_eval,
        dpath_trial,
        partitions,
        bucket_partition,
        bucket_comp_keys_full_set,
        fontsize_axes,
        fontsize_ticks,
        fontsize_legend,
        subplot_border_width,
        figsize,
        height_ratios,
        partition_metric_group="per_class",
        full_set=True,
        retrieval_ylabel="Full-Set Macro mAP Scores",
        accuracy_ylabel="Full-Set I2T Per-Class Accuracy",
        nshot_accuracy_ylabel="Full-Set n-shot Per-Class\nAccuracy (ID)",
        plot_title=f"Train Metrics (Macro Full-Set){title_suffix}",
        output_filename="full_macro.png",
    )

def plot_composite_metrics(
    data_epoch,
    data_eval,
    x_train,
    x_eval,
    dpath_trial,
    partitions,
    bucket_partition,
    bucket_comp_keys,
    fontsize_axes,
    fontsize_ticks,
    fontsize_legend,
    subplot_border_width,
    figsize,
    height_ratios,
    partition_metric_group,
    full_set,
    retrieval_ylabel,
    accuracy_ylabel,
    nshot_accuracy_ylabel,
    plot_title,
    output_filename,
):
    fig = plt.figure(figsize=figsize)
    gs = gridspec.GridSpec(len(height_ratios), 1, height_ratios=height_ratios, hspace=0)

    ax0 = fig.add_subplot(gs[0, 0])

    id_partition = "id" if "id" in partitions else None
    ood_partition = "ood" if "ood" in partitions else None
    retrieval_specs = (
        ("i2t", "I2T", "blue"),
        ("i2i", "I2I", "red"),
        ("t2i", "T2I", "green"),
    )
    set_key = "full_set" if full_set else "closed_set"
    style_specs = (
        (id_partition, "ID", "-"),
        (ood_partition, "OOD", "--"),
    )
    for partition, partition_label, linestyle in style_specs:
        if partition is None:
            continue
        partition_group_scores = data_eval.get("scores", {}).get(set_key, {}).get(partition_metric_group, {}).get(partition, {})
        partition_group_scores = partition_group_scores.get("map", {})
        for metric_name, metric_label, color in retrieval_specs:
            if metric_name in partition_group_scores:
                ax0.plot(
                    x_eval,
                    partition_group_scores[metric_name],
                    label=f"{partition_label} {metric_label}",
                    color=color,
                    linestyle=linestyle,
                )

    ax0.set_ylabel(retrieval_ylabel, fontsize=fontsize_axes, fontweight="bold")
    ax0.set_ylim(0, 1)
    if partitions:
        ax0.legend(loc="lower right", fontsize=fontsize_legend)
    ax0.grid(True)
    ax0.tick_params(labelbottom=False, labelsize=fontsize_ticks)

    ax1 = fig.add_subplot(gs[1, 0], sharex=ax0)
    is_macro_plot = partition_metric_group == "per_class"
    id_mode_scores = data_eval.get("scores", {}).get(set_key, {}).get(partition_metric_group, {}).get(bucket_partition, {})
    nshot_key = "n-shot"
    if is_macro_plot:
        nshot_ylabel = "Full-Set n-shot Macro mAP (ID)" if full_set else "n-shot Macro mAP (ID)"
    else:
        nshot_ylabel = "Full-Set n-shot mAP (ID)" if full_set else "n-shot mAP (ID)"
    comp_nshot = id_mode_scores.get("map", {}).get(nshot_key, {})
    if bucket_comp_keys:
        for key in reversed(bucket_comp_keys):
            label = key
            maybe_plot(ax1, x_eval, comp_nshot, key, label, linestyle=":")
        if comp_nshot:
            ax1.legend(loc="lower right", fontsize=fontsize_legend)
    ax1.set_ylabel(nshot_ylabel, fontsize=fontsize_axes, fontweight="bold")
    ax1.set_ylim(0, 1)
    ax1.grid(True)
    ax1.tick_params(labelbottom=False, labelsize=fontsize_ticks)

    ax2 = fig.add_subplot(gs[2, 0], sharex=ax0)
    for partition in partitions:
        partition_group_scores = data_eval.get("scores", {}).get(set_key, {}).get(partition_metric_group, {}).get(partition, {})
        partition_group_scores = partition_group_scores.get("acc", {})
        if "i2t" in partition_group_scores:
            ax2.plot(
                x_eval,
                partition_group_scores["i2t"],
                label="-".join([s.upper() if i == 0 else s.title() for i, s in enumerate(partition.split("_"))]),
            )
    ax2.set_ylabel(accuracy_ylabel, fontsize=fontsize_axes, fontweight="bold")
    ax2.set_ylim(0, 1)
    if partitions:
        ax2.legend(loc="lower right", fontsize=fontsize_legend)
    ax2.grid(True)
    ax2.tick_params(labelbottom=False, labelsize=fontsize_ticks)

    ax3 = fig.add_subplot(gs[3, 0], sharex=ax0)
    comp_nshot_acc = id_mode_scores.get("acc", {}).get("n-shot", {})
    if bucket_comp_keys:
        for key in reversed(bucket_comp_keys):
            maybe_plot(ax3, x_eval, comp_nshot_acc, key, key, linestyle=":")
        if comp_nshot_acc:
            ax3.legend(loc="lower right", fontsize=fontsize_legend)
    ax3.set_ylabel(nshot_accuracy_ylabel, fontsize=fontsize_axes, fontweight="bold")
    ax3.set_ylim(0, 1)
    ax3.grid(True)
    ax3.tick_params(labelbottom=False, labelsize=fontsize_ticks)

    ax4 = fig.add_subplot(gs[4, 0], sharex=ax0)
    if len(data_epoch.get("loss_train", [])) == len(x_train):
        ax4.plot(x_train, data_epoch["loss_train"], label="Train Loss")
    if len(data_epoch.get("loss_raw_train", [])) == len(x_train):
        ax4.plot(x_train, data_epoch["loss_raw_train"], label="Train Loss (Raw)")
    for partition in partitions:
        if len(data_eval.get("loss_raw", {}).get(partition, [])) == len(x_eval):
            ax4.plot(
                x_eval,
                data_eval["loss_raw"][partition],
                label=f'{"-".join([s.upper() if i == 0 else s.title() for i, s in enumerate(partition.split("_"))])} Val Loss',
            )
    ax4.set_ylabel("Loss", fontsize=fontsize_axes, fontweight="bold")
    ax4.set_yscale("log")
    ax4.minorticks_on()
    ax4.grid(which="minor", axis="y")
    ax4.legend(loc="upper right", fontsize=fontsize_legend)
    ax4.grid(True)
    ax4.tick_params(labelbottom=False, labelsize=fontsize_ticks)

    ax5 = fig.add_subplot(gs[5, 0], sharex=ax0)
    if len(data_epoch.get("grad_norm_model", [])) == len(x_train):
        ax5.plot(x_train, data_epoch["grad_norm_model"], color="green")
    ax5.set_ylabel("Grad Norm", fontsize=fontsize_axes, fontweight="bold")
    ax5.set_yscale("log")
    ax5.minorticks_on()
    ax5.grid(which="minor", axis="y")
    ax5.grid(True)
    ax5.tick_params(labelbottom=False, labelsize=fontsize_ticks)

    ax6 = fig.add_subplot(gs[6, 0], sharex=ax0)
    color_sim = "tab:blue"
    color_targ = "tab:orange"
    # targets first, then similarity; min/max solid, mean dashed, median dotted
    for stat_prefix, stat_color in (("targ", color_targ), ("sim", color_sim)):
        for stat_name, stat_linestyle in (("min", "-"), ("max", "-"), ("mean", "--"), ("median", ":")):
            stat_key = f"{stat_prefix}_{stat_name}"
            if len(data_epoch.get(stat_key, [])) == len(x_train):
                ax6.plot(x_train, data_epoch[stat_key], color=stat_color, linestyle=stat_linestyle, linewidth=1.0)
    ax6.set_ylabel("Similarity / Target", fontsize=fontsize_axes, fontweight="bold")
    ax6.set_ylim(-1.0, 1.0)
    ax6.legend(
        handles=[
            Line2D([0], [0], color=color_sim, lw=1.0, label="Similarity"),
            Line2D([0], [0], color=color_targ, lw=1.0, label="Target"),
            Line2D([0], [0], color="gray", lw=1.0, linestyle="-", label="Min/Max"),
            Line2D([0], [0], color="gray", lw=1.0, linestyle="--", label="Mean"),
            Line2D([0], [0], color="gray", lw=1.0, linestyle=":", label="Median"),
        ],
        loc="upper center",
        ncol=5,
        fontsize=fontsize_legend,
    )
    ax6.grid(True)
    ax6.tick_params(labelbottom=False, labelsize=fontsize_ticks)

    ax7 = fig.add_subplot(gs[7, 0], sharex=ax0)
    if len(data_epoch.get("lr", [])) == len(x_train):
        ax7.plot(x_train, data_epoch["lr"])
    ax7.set_ylabel("Learning Rate", fontsize=fontsize_axes, fontweight="bold")
    ax7.ticklabel_format(style="sci", axis="y", scilimits=(0, 0))
    ax7.yaxis.set_offset_position("right")
    ax7.yaxis.set_major_formatter(FormatStrFormatter("%.1e"))
    ax7.yaxis.get_offset_text().set_visible(False)
    ax7.set_xlabel("Samples Seen (M)", fontsize=fontsize_axes, fontweight="bold")
    ax7.xaxis.set_major_formatter(FuncFormatter(_samples_seen_tick_formatter))
    ax7.grid(True)
    ax7.tick_params(labelsize=fontsize_ticks)

    for ax in (ax0, ax1, ax2, ax3, ax4, ax5, ax6, ax7):
        ax.label_outer()

    for idx_ax, ax in enumerate((ax0, ax1, ax2, ax3, ax4, ax5, ax6, ax7)):
        for spine in ax.spines.values():
            spine.set_linewidth(subplot_border_width)
            spine.set_edgecolor("black")
        if idx_ax % 2 == 1:
            ax.yaxis.set_label_position("right")
            ax.yaxis.tick_right()

    fig.suptitle(plot_title, fontweight="bold", y=0.98, fontsize=20)
    plt.subplots_adjust(hspace=0)
    plt.tight_layout()
    plots_dir = dpath_trial / "learning_curves"
    plots_dir.mkdir(parents=True, exist_ok=True)
    fig.savefig(plots_dir / output_filename, dpi=300)
    plt.close(fig)

def maybe_plot(ax, x, data, key, label, **kwargs):
    """
    Helper for plot_metrics() (N-Shot Composites)
    """
    if key in data and len(data[key]) > 0:
        ax.plot(x, data[key], label=label, **kwargs)
